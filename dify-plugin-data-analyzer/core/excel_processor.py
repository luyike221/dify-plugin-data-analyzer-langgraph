"""
Excelæ™ºèƒ½å¤„ç†æ¨¡å—
æ”¯æŒï¼š
1. è‡ªåŠ¨è·³è¿‡æ— æ•ˆè¡Œï¼ˆæ³¨é‡Šã€æ ‡é¢˜ç­‰ï¼‰
2. å•è¡¨å¤´/å¤šè¡¨å¤´è‡ªåŠ¨è¯†åˆ«
3. å¯é€‰è°ƒç”¨LLMè¿›è¡Œæ™ºèƒ½åˆ†æ
4. åˆå¹¶å•å…ƒæ ¼å¤„ç†
5. åˆ—ç»“æ„å…ƒæ•°æ®ç”Ÿæˆ
"""

import pandas as pd
import json
import re
import os
import requests
import logging
from openpyxl import load_workbook
from typing import Tuple, List, Dict, Optional, Any
from collections import defaultdict
from dataclasses import dataclass, asdict, field
from pathlib import Path

# é…ç½®æ—¥å¿—
logger = logging.getLogger(__name__)

# å¯¼å…¥é…ç½®ï¼ˆé¿å…å¾ªç¯å¯¼å…¥ï¼Œä½¿ç”¨å»¶è¿Ÿå¯¼å…¥ï¼‰
try:
    from .config import EXCEL_LLM_API_KEY, EXCEL_LLM_BASE_URL, EXCEL_LLM_MODEL
except ImportError:
    # å¦‚æœæ— æ³•å¯¼å…¥ï¼Œä½¿ç”¨ç¯å¢ƒå˜é‡
    EXCEL_LLM_API_KEY = os.environ.get("EXCEL_LLM_API_KEY", "")
    EXCEL_LLM_BASE_URL = os.environ.get("EXCEL_LLM_BASE_URL", "https://api.openai.com/v1/chat/completions")
    EXCEL_LLM_MODEL = os.environ.get("EXCEL_LLM_MODEL", "gpt-4o-mini")


@dataclass
class HeaderAnalysis:
    """è¡¨å¤´åˆ†æç»“æœ"""
    skip_rows: int          # éœ€è¦è·³è¿‡çš„æ— æ•ˆè¡Œæ•°
    header_rows: int        # è¡¨å¤´å ç”¨çš„è¡Œæ•°
    header_type: str        # 'single' æˆ– 'multi'
    data_start_row: int     # æ•°æ®å¼€å§‹è¡Œï¼ˆ1-indexedï¼‰
    confidence: str         # ç½®ä¿¡åº¦: high/medium/low
    reason: str             # åˆ†æåŸå› è¯´æ˜
    valid_cols: Optional[List[int]] = None  # æœ‰æ•ˆåˆ—çš„ç´¢å¼•åˆ—è¡¨ï¼ˆ1-indexedï¼‰ï¼ŒNoneè¡¨ç¤ºæ‰€æœ‰åˆ—éƒ½æœ‰æ•ˆ
    
    def to_dict(self) -> Dict[str, Any]:
        """è½¬æ¢ä¸ºå­—å…¸"""
        result = asdict(self)
        if result.get('valid_cols') is None:
            result['valid_cols'] = None
        return result


@dataclass
class ExcelProcessResult:
    """Excelå¤„ç†ç»“æœ"""
    success: bool
    header_analysis: Optional[HeaderAnalysis]
    processed_file_path: Optional[str]      # å¤„ç†åçš„CSVæ–‡ä»¶è·¯å¾„
    metadata_file_path: Optional[str]       # å…ƒæ•°æ®JSONæ–‡ä»¶è·¯å¾„
    column_names: List[str]                 # åˆ—ååˆ—è¡¨
    column_metadata: Dict[str, Dict]        # åˆ—ç»“æ„å…ƒæ•°æ®
    row_count: int                          # æ•°æ®è¡Œæ•°
    error_message: Optional[str]            # é”™è¯¯ä¿¡æ¯
    
    def to_dict(self) -> Dict[str, Any]:
        """è½¬æ¢ä¸ºå­—å…¸"""
        return {
            "success": self.success,
            "header_analysis": self.header_analysis.to_dict() if self.header_analysis else None,
            "processed_file_path": self.processed_file_path,
            "metadata_file_path": self.metadata_file_path,
            "column_names": self.column_names,
            "column_metadata": self.column_metadata,
            "row_count": self.row_count,
            "error_message": self.error_message
        }


class SmartHeaderProcessor:
    """æ™ºèƒ½è¡¨å¤´å¤„ç†å™¨"""
    
    def __init__(self, filepath: str, sheet_name: str = None):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.wb = load_workbook(filepath, data_only=True)
        self.ws = self.wb[sheet_name] if sheet_name else self.wb.active
        self.merged_cells_map = self._build_merged_cells_map()
    
    def _build_merged_cells_map(self) -> Dict[Tuple[int, int], str]:
        """æ„å»ºåˆå¹¶å•å…ƒæ ¼æ˜ å°„"""
        merged_map = {}
        for merged_range in self.ws.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            value = self.ws.cell(min_row, min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_map[(row, col)] = value
        return merged_map
    
    def get_cell_value(self, row: int, col: int) -> Any:
        """è·å–å•å…ƒæ ¼å€¼ï¼Œå¤„ç†åˆå¹¶å•å…ƒæ ¼"""
        if (row, col) in self.merged_cells_map:
            return self.merged_cells_map[(row, col)]
        return self.ws.cell(row, col).value
    
    def get_preview_data(self, max_rows: int = 15, max_cols: int = 10) -> List[List[Any]]:
        """è·å–é¢„è§ˆæ•°æ®ç”¨äºåˆ†æ"""
        actual_max_col = min(self.ws.max_column, max_cols)
        actual_max_row = min(self.ws.max_row, max_rows)
        
        data = []
        for row in range(1, actual_max_row + 1):
            row_data = []
            for col in range(1, actual_max_col + 1):
                value = self.get_cell_value(row, col)
                # è½¬æ¢ä¸ºå­—ç¬¦ä¸²ä¾¿äºåˆ†æ
                if value is None:
                    row_data.append("")
                elif isinstance(value, (int, float)):
                    row_data.append(f"[æ•°å€¼:{value}]")
                else:
                    row_data.append(str(value)[:50])  # æˆªæ–­è¿‡é•¿å†…å®¹
            data.append(row_data)
        return data
    
    def get_merged_info(self) -> List[Dict]:
        """è·å–åˆå¹¶å•å…ƒæ ¼ä¿¡æ¯"""
        merged_info = []
        for merged_range in self.ws.merged_cells.ranges:
            if merged_range.min_row <= 10:  # åªå…³æ³¨å‰10è¡Œ
                merged_info.append({
                    'range': str(merged_range),
                    'rows': f"{merged_range.min_row}-{merged_range.max_row}",
                    'cols': f"{merged_range.min_col}-{merged_range.max_col}",
                    'value': str(self.ws.cell(merged_range.min_row, merged_range.min_col).value)[:30]
                })
        return merged_info
    
    def validate_with_llm(self, rule_analysis: HeaderAnalysis, 
                         llm_api_key: Optional[str] = None,
                         llm_base_url: Optional[str] = None,
                         llm_model: Optional[str] = None) -> HeaderAnalysis:
        """
        ä½¿ç”¨LLMéªŒè¯è§„åˆ™åˆ†æçš„ç»“æœ
        
        å‚æ•°:
            rule_analysis: è§„åˆ™åˆ†æçš„ç»“æœ
            llm_api_key: LLM APIå¯†é’¥ï¼ˆå¯é€‰ï¼‰
            llm_base_url: LLM APIåœ°å€ï¼ˆå¯é€‰ï¼‰
            llm_model: LLMæ¨¡å‹åç§°ï¼ˆå¯é€‰ï¼‰
        
        è¿”å›:
            éªŒè¯åçš„åˆ†æç»“æœï¼ˆå¦‚æœLLMéªŒè¯å¤±è´¥ï¼Œè¿”å›åŸè§„åˆ™åˆ†æç»“æœï¼‰
        """
        preview_data = self.get_preview_data()
        merged_info = self.get_merged_info()
        
        # æ„å»ºéªŒè¯æç¤ºè¯
        prompt = self._build_validation_prompt(preview_data, merged_info, rule_analysis)
        
        # è°ƒç”¨LLMï¼ˆä½¿ç”¨ä¼ å…¥çš„é…ç½®æˆ–ä»å…¨å±€é…ç½®è¯»å–ï¼‰
        result = self._call_llm(prompt, llm_api_key, llm_base_url, llm_model)
        
        # è§£æLLMéªŒè¯ç»“æœ
        validated = self._parse_validation_response(result, rule_analysis)
        
        return validated
    
    def _build_validation_prompt(self, preview_data: List[List], merged_info: List[Dict], 
                                rule_analysis: HeaderAnalysis) -> str:
        """æ„å»ºLLMéªŒè¯æç¤ºè¯"""
        # æ ¼å¼åŒ–é¢„è§ˆæ•°æ®ä¸ºè¡¨æ ¼å½¢å¼
        table_str = "è¡Œå· | å†…å®¹\n" + "-" * 50 + "\n"
        for i, row in enumerate(preview_data, 1):
            row_str = " | ".join(str(cell)[:20] for cell in row[:8])
            table_str += f"  {i}  | {row_str}\n"
        
        # æ ¼å¼åŒ–åˆå¹¶å•å…ƒæ ¼ä¿¡æ¯
        merged_str = "æ— " if not merged_info else "\n".join(
            f"  - {m['range']}: '{m['value']}'" for m in merged_info[:5]
        )
        
        prompt = f"""è¯·éªŒè¯ä»¥ä¸‹Excelè¡¨æ ¼çš„è§„åˆ™åˆ†æç»“æœæ˜¯å¦æ­£ç¡®ã€‚

ã€è¡¨æ ¼é¢„è§ˆã€‘ï¼ˆå‰15è¡Œï¼Œ[æ•°å€¼:xxx]è¡¨ç¤ºæ•°å€¼ç±»å‹ï¼‰
{table_str}

ã€åˆå¹¶å•å…ƒæ ¼ã€‘
{merged_str}

ã€è§„åˆ™åˆ†æç»“æœã€‘
- è·³è¿‡è¡Œæ•°: {rule_analysis.skip_rows}
- è¡¨å¤´è¡Œæ•°: {rule_analysis.header_rows}
- è¡¨å¤´ç±»å‹: {rule_analysis.header_type}
- æ•°æ®èµ·å§‹è¡Œ: {rule_analysis.data_start_row}
- åˆ†æåŸå› : {rule_analysis.reason}

è¯·éªŒè¯è¿™ä¸ªç»“æœæ˜¯å¦åˆç†ï¼Œå¹¶ä»¥JSONæ ¼å¼è¿”å›ï¼š
{{
    "is_valid": <trueæˆ–falseï¼Œè¡¨ç¤ºç»“æœæ˜¯å¦åˆç†>,
    "confidence": "<high/medium/low>",
    "suggestions": {{
        "skip_rows": <å»ºè®®çš„è·³è¿‡è¡Œæ•°ï¼Œå¦‚æœåˆç†åˆ™ä¸è§„åˆ™åˆ†æç›¸åŒ>,
        "header_rows": <å»ºè®®çš„è¡¨å¤´è¡Œæ•°ï¼Œå¦‚æœåˆç†åˆ™ä¸è§„åˆ™åˆ†æç›¸åŒ>,
        "header_type": "<singleæˆ–multi>",
        "data_start_row": <å»ºè®®çš„æ•°æ®èµ·å§‹è¡Œï¼Œå¦‚æœåˆç†åˆ™ä¸è§„åˆ™åˆ†æç›¸åŒ>
    }},
    "reason": "<éªŒè¯è¯´æ˜ï¼šå¦‚æœåˆç†ï¼Œè¯´æ˜ä¸ºä»€ä¹ˆï¼›å¦‚æœä¸åˆç†ï¼ŒæŒ‡å‡ºé—®é¢˜å¹¶ç»™å‡ºå»ºè®®>"
}}

éªŒè¯è¦ç‚¹ï¼š
- æ£€æŸ¥è·³è¿‡çš„è¡Œæ˜¯å¦çœŸçš„æ˜¯æ— æ•ˆè¡Œï¼ˆæ ‡é¢˜ã€æ³¨é‡Šç­‰ï¼‰
- æ£€æŸ¥è¡¨å¤´è¡Œæ•°æ˜¯å¦æ­£ç¡®ï¼ˆæ˜¯å¦é—æ¼äº†å¤šçº§è¡¨å¤´ï¼‰
- æ£€æŸ¥æ•°æ®èµ·å§‹è¡Œæ˜¯å¦å‡†ç¡®ï¼ˆæ˜¯å¦æŠŠè¡¨å¤´è¡Œè¯¯åˆ¤ä¸ºæ•°æ®è¡Œï¼‰
- å¦‚æœè§„åˆ™åˆ†æç»“æœåˆç†ï¼Œä¿æŒåŸç»“æœï¼›å¦‚æœä¸åˆç†ï¼Œç»™å‡ºä¿®æ­£å»ºè®®
- åªè¿”å›JSONï¼Œä¸è¦å…¶ä»–å†…å®¹"""
        
        return prompt
    
    def _call_llm(self, prompt: str, llm_api_key: Optional[str] = None, 
                  llm_base_url: Optional[str] = None, llm_model: Optional[str] = None) -> str:
        """è°ƒç”¨LLM APIï¼ˆæ”¯æŒOpenAIå…¼å®¹æ¥å£ï¼‰
        
        å‚æ•°:
            prompt: æç¤ºè¯
            llm_api_key: LLM APIå¯†é’¥ï¼ˆå¯é€‰ï¼Œå¦‚æœä¸æä¾›åˆ™ä»é…ç½®è¯»å–ï¼‰
            llm_base_url: LLM APIåœ°å€ï¼ˆå¯é€‰ï¼Œå¦‚æœä¸æä¾›åˆ™ä»é…ç½®è¯»å–ï¼‰
            llm_model: LLMæ¨¡å‹åç§°ï¼ˆå¯é€‰ï¼Œå¦‚æœä¸æä¾›åˆ™ä»é…ç½®è¯»å–ï¼‰
        """
        # ä¼˜å…ˆä½¿ç”¨ä¼ å…¥çš„å‚æ•°ï¼Œå¦åˆ™ä»é…ç½®è¯»å–
        api_key = llm_api_key if llm_api_key is not None else EXCEL_LLM_API_KEY
        base_url = llm_base_url if llm_base_url is not None else EXCEL_LLM_BASE_URL
        model = llm_model if llm_model is not None else EXCEL_LLM_MODEL
        
        logger.info("=" * 60)
        logger.info("ğŸ¤– è°ƒç”¨ LLM API è¿›è¡Œè¡¨å¤´éªŒè¯")
        logger.info(f"ğŸ”— EXCEL_LLM_BASE_URL: {base_url}")
        logger.info(f"ğŸ“Œ æ¨¡å‹: {model}")
        logger.info(f"ğŸ”‘ API Key: {'å·²é…ç½®' if api_key else 'æœªé…ç½®'}")
        
        if not api_key:
            logger.warning("âš ï¸ æœªé…ç½® LLM API Keyï¼Œè·³è¿‡ LLM éªŒè¯")
            return None
            
        url = base_url
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        payload = {
            "model": model,
            "max_tokens": 500,
            "messages": [{"role": "user", "content": prompt}]
        }
        
        logger.info(f"ğŸ“¡ å‘é€ LLM API è¯·æ±‚åˆ°: {url}")
        logger.info(f"ğŸ“ æç¤ºè¯é•¿åº¦: {len(prompt)} å­—ç¬¦")
        
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=30)
            response.raise_for_status()
            result = response.json()
            llm_response = result['choices'][0]['message']['content']
            
            logger.info("âœ… LLM API è°ƒç”¨æˆåŠŸ")
            logger.info("=" * 60)
            logger.info("ğŸ“ LLM å“åº”å†…å®¹:")
            logger.info("=" * 60)
            logger.info(llm_response)
            logger.info("=" * 60)
            
            return llm_response
        except Exception as e:
            logger.error(f"âŒ LLMè°ƒç”¨å¤±è´¥: {e}")
            logger.debug("å¼‚å¸¸è¯¦æƒ…:", exc_info=True)
            return None
    
    def _parse_validation_response(self, response: str, rule_analysis: HeaderAnalysis) -> HeaderAnalysis:
        """è§£æLLMéªŒè¯ç»“æœ"""
        if not response:
            # LLMè°ƒç”¨å¤±è´¥ï¼Œè¿”å›åŸè§„åˆ™åˆ†æç»“æœ
            return rule_analysis
        
        try:
            # æå–JSONéƒ¨åˆ†ï¼ˆæ”¯æŒåµŒå¥—JSONï¼‰
            # å…ˆå°è¯•æ‰¾åˆ°ç¬¬ä¸€ä¸ª { åˆ°æœ€åä¸€ä¸ª } ä¹‹é—´çš„å†…å®¹
            start_idx = response.find('{')
            end_idx = response.rfind('}')
            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                json_str = response[start_idx:end_idx + 1]
                data = json.loads(json_str)
            else:
                # å¦‚æœæ‰¾ä¸åˆ°å®Œæ•´çš„JSONï¼Œå°è¯•ç”¨æ­£åˆ™åŒ¹é…
                json_match = re.search(r'\{.*\}', response, re.DOTALL)
                if not json_match:
                    raise ValueError("æœªæ‰¾åˆ°JSONæ ¼å¼çš„å“åº”")
                data = json.loads(json_match.group())
            
            is_valid = data.get('is_valid', True)
            suggestions = data.get('suggestions', {})
            
            if is_valid:
                # LLMè®¤ä¸ºè§„åˆ™åˆ†æç»“æœåˆç†ï¼Œä¿æŒåŸç»“æœä½†æ›´æ–°ç½®ä¿¡åº¦å’ŒåŸå› 
                return HeaderAnalysis(
                    skip_rows=rule_analysis.skip_rows,
                    header_rows=rule_analysis.header_rows,
                    header_type=rule_analysis.header_type,
                    data_start_row=rule_analysis.data_start_row,
                    confidence=data.get('confidence', 'high'),  # LLMéªŒè¯é€šè¿‡ï¼Œç½®ä¿¡åº¦æå‡
                    reason=f"è§„åˆ™åˆ†æ+LLMéªŒè¯: {data.get('reason', 'éªŒè¯é€šè¿‡')}",
                    valid_cols=rule_analysis.valid_cols  # ä¿æŒåŸæœ‰çš„åˆ—è¿‡æ»¤ç»“æœ
                )
            else:
                # LLMè®¤ä¸ºä¸åˆç†ï¼Œä½¿ç”¨LLMçš„å»ºè®®
                # æ³¨æ„ï¼šLLMå¯èƒ½å»ºè®®ä¿®æ”¹è¡¨å¤´è¡Œæ•°ï¼Œä½†åˆ—è¿‡æ»¤ç»“æœä»ç„¶ä¿ç•™
                return HeaderAnalysis(
                    skip_rows=suggestions.get('skip_rows', rule_analysis.skip_rows),
                    header_rows=suggestions.get('header_rows', rule_analysis.header_rows),
                    header_type=suggestions.get('header_type', rule_analysis.header_type),
                    data_start_row=suggestions.get('data_start_row', rule_analysis.data_start_row),
                    confidence=data.get('confidence', 'medium'),
                    reason=f"è§„åˆ™åˆ†æ+LLMä¿®æ­£: {data.get('reason', 'LLMå»ºè®®ä¿®æ­£')}",
                    valid_cols=rule_analysis.valid_cols  # ä¿æŒåŸæœ‰çš„åˆ—è¿‡æ»¤ç»“æœ
                )
        except (json.JSONDecodeError, KeyError, ValueError) as e:
            print(f"è§£æLLMéªŒè¯å“åº”å¤±è´¥: {e}ï¼Œä½¿ç”¨åŸè§„åˆ™åˆ†æç»“æœ")
        
        # è§£æå¤±è´¥ï¼Œè¿”å›åŸè§„åˆ™åˆ†æç»“æœ
        return rule_analysis
    
    def analyze_with_rules(self) -> HeaderAnalysis:
        """åŸºäºè§„åˆ™çš„åˆ†æï¼ˆä½œä¸ºLLMçš„é™çº§æ–¹æ¡ˆï¼‰"""
        max_col = self.ws.max_column
        skip_rows = 0
        header_rows = 1
        
        # æ£€æµ‹éœ€è¦è·³è¿‡çš„è¡Œ
        for row in range(1, min(6, self.ws.max_row + 1)):
            row_values = [self.get_cell_value(row, col) for col in range(1, max_col + 1)]
            non_empty = sum(1 for v in row_values if v is not None)
            
            # å¦‚æœåªæœ‰å¾ˆå°‘çš„éç©ºå•å…ƒæ ¼ï¼Œå¯èƒ½æ˜¯æ ‡é¢˜è¡Œ
            if non_empty <= 2 and non_empty < max_col * 0.3:
                skip_rows = row
            else:
                break
        
        # æ£€æµ‹è¡¨å¤´è¡Œæ•°
        header_start = skip_rows + 1
        
        # æ£€æŸ¥åˆå¹¶å•å…ƒæ ¼
        max_merged_row = 0
        for merged_range in self.ws.merged_cells.ranges:
            if merged_range.min_row > skip_rows:
                if merged_range.max_row > max_merged_row:
                    max_merged_row = merged_range.max_row
        
        if max_merged_row > header_start:
            header_rows = max_merged_row - skip_rows
        
        # æ£€æµ‹æ•°æ®è¡Œå¼€å§‹ä½ç½®
        data_start = skip_rows + header_rows + 1
        for row in range(header_start, min(skip_rows + 10, self.ws.max_row + 1)):
            row_values = [self.get_cell_value(row, col) for col in range(1, max_col + 1)]
            non_empty = sum(1 for v in row_values if v is not None)
            numeric = sum(1 for v in row_values if isinstance(v, (int, float)) and not isinstance(v, bool))
            
            if non_empty > 0 and numeric / max(non_empty, 1) > 0.4:
                data_start = row
                header_rows = row - skip_rows - 1
                break
        
        header_type = 'multi' if header_rows > 1 else 'single'
        
        # æ³¨æ„ï¼šåˆ—æ£€æµ‹åœ¨LLMéªŒè¯å®Œæˆåè¿›è¡Œï¼Œè¿™é‡Œä¸è¿›è¡Œåˆ—æ£€æµ‹
        return HeaderAnalysis(
            skip_rows=skip_rows,
            header_rows=max(1, header_rows),
            header_type=header_type,
            data_start_row=data_start,
            confidence='medium',
            reason='åŸºäºè§„åˆ™åˆ†æ',
            valid_cols=None  # åˆ—æ£€æµ‹åœ¨LLMéªŒè¯å®Œæˆåè¿›è¡Œ
        )
    
    def _detect_valid_columns(self, skip_rows: int, header_rows: int, data_start_row: int) -> List[int]:
        """
        æ£€æµ‹æœ‰æ•ˆåˆ—ï¼ˆè¿‡æ»¤æ— æ•ˆåˆ—ï¼‰
        
        æ— æ•ˆåˆ—çš„åˆ¤æ–­æ ‡å‡†ï¼š
        1. è¡¨å¤´åŒºåŸŸå®Œå…¨ä¸ºç©º
        2. æ•°æ®åŒºåŸŸå®Œå…¨ä¸ºç©ºæˆ–æ²¡æœ‰æ•°å€¼æ•°æ®
        
        è¿”å›: æœ‰æ•ˆåˆ—çš„ç´¢å¼•åˆ—è¡¨ï¼ˆ1-indexedï¼‰
        """
        max_col = self.ws.max_column
        header_start = skip_rows + 1
        header_end = skip_rows + header_rows
        valid_cols = []
        
        logger.info("ğŸ” å¼€å§‹æ£€æµ‹æ— æ•ˆåˆ—...")
        
        for col in range(1, max_col + 1):
            # æ£€æŸ¥è¡¨å¤´åŒºåŸŸæ˜¯å¦æœ‰å†…å®¹
            has_header = False
            for row in range(header_start, header_end + 1):
                value = self.get_cell_value(row, col)
                if value is not None and str(value).strip():
                    has_header = True
                    break
            
            # æ£€æŸ¥æ•°æ®åŒºåŸŸæ˜¯å¦æœ‰æ•°å€¼æ•°æ®
            has_data = False
            numeric_count = 0
            total_count = 0
            for row in range(data_start_row, min(data_start_row + 10, self.ws.max_row + 1)):
                value = self.ws.cell(row, col).value
                if value is not None:
                    total_count += 1
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        numeric_count += 1
                        has_data = True
            
            # å¦‚æœè¡¨å¤´æœ‰å†…å®¹æˆ–æ•°æ®åŒºåŸŸæœ‰æ•°å€¼ï¼Œåˆ™è®¤ä¸ºæ˜¯æœ‰æ•ˆåˆ—
            if has_header or has_data:
                valid_cols.append(col)
                logger.debug(f"âœ… åˆ— {col}: æœ‰æ•ˆ (è¡¨å¤´: {has_header}, æ•°æ®: {has_data}, æ•°å€¼: {numeric_count}/{total_count})")
            else:
                logger.info(f"âŒ åˆ— {col}: æ— æ•ˆ (è¡¨å¤´ä¸ºç©ºä¸”æ•°æ®ä¸ºç©º)")
        
        logger.info(f"ğŸ“Š åˆ—è¿‡æ»¤ç»“æœ: æ€»åˆ—æ•° {max_col}, æœ‰æ•ˆåˆ—æ•° {len(valid_cols)}, æ— æ•ˆåˆ—æ•° {max_col - len(valid_cols)}")
        
        # å¦‚æœæ‰€æœ‰åˆ—éƒ½æœ‰æ•ˆï¼Œè¿”å›Noneï¼ˆè¡¨ç¤ºä¸éœ€è¦è¿‡æ»¤ï¼‰
        if len(valid_cols) == max_col:
            return None
        
        return valid_cols
    
    def extract_headers(self, analysis: HeaderAnalysis) -> Tuple[List[str], Dict[str, Dict]]:
        """
        æ ¹æ®åˆ†æç»“æœæå–è¡¨å¤´
        è¿”å›: (åˆ—ååˆ—è¡¨, åˆ—ç»“æ„å…ƒæ•°æ®)
        """
        max_col = self.ws.max_column
        header_start = analysis.skip_rows + 1
        header_end = analysis.skip_rows + analysis.header_rows
        
        # ç¡®å®šè¦å¤„ç†çš„åˆ—ï¼ˆå¦‚æœæŒ‡å®šäº†æœ‰æ•ˆåˆ—ï¼Œåªå¤„ç†æœ‰æ•ˆåˆ—ï¼‰
        cols_to_process = analysis.valid_cols if analysis.valid_cols is not None else list(range(1, max_col + 1))
        
        logger.info(f"ğŸ“‹ æå–è¡¨å¤´: å¤„ç† {len(cols_to_process)} åˆ—")
        
        column_metadata = {}
        
        if analysis.header_type == 'single':
            # å•è¡¨å¤´
            headers = []
            for col in cols_to_process:
                value = self.get_cell_value(header_start, col)
                col_name = str(value) if value else f'Column_{col}'
                headers.append(col_name)
                column_metadata[col_name] = {"level1": col_name}
            
            headers = self._handle_duplicate_names(headers)
            # æ›´æ–°å…ƒæ•°æ®çš„key
            column_metadata = {h: {"level1": h} for h in headers}
            return headers, column_metadata
        
        else:
            # å¤šè¡¨å¤´ï¼šå±•å¹³
            column_headers = []
            for col in cols_to_process:
                parts = []
                levels = {}
                for row_idx, row in enumerate(range(header_start, header_end + 1), 1):
                    value = self.get_cell_value(row, col)
                    if value is not None:
                        part = str(value).strip()
                        parts.append(part)
                        levels[f"level{row_idx}"] = part
                
                # å»é‡è¿ç»­ç›¸åŒå€¼
                unique_parts = []
                for p in parts:
                    if not unique_parts or p != unique_parts[-1]:
                        unique_parts.append(p)
                
                col_name = '_'.join(unique_parts) if unique_parts else f'Column_{col}'
                column_headers.append(col_name)
                column_metadata[col_name] = levels
            
            column_headers = self._handle_duplicate_names(column_headers)
            
            # é‡æ–°æ˜ å°„å…ƒæ•°æ®
            new_metadata = {}
            for i, header in enumerate(column_headers):
                original_name = '_'.join(unique_parts) if (unique_parts := list(column_metadata.values())[i].values()) else f'Column_{i+1}'
                new_metadata[header] = list(column_metadata.values())[i]
            
            return column_headers, new_metadata
    
    def _handle_duplicate_names(self, names: List[str]) -> List[str]:
        """å¤„ç†é‡å¤åˆ—å"""
        counts = defaultdict(int)
        result = []
        for name in names:
            if counts[name] > 0:
                result.append(f"{name}_{counts[name]}")
            else:
                result.append(name)
            counts[name] += 1
        return result
    
    def to_dataframe(self, analysis: HeaderAnalysis = None, use_llm_validate: bool = False,
                    llm_api_key: Optional[str] = None,
                    llm_base_url: Optional[str] = None,
                    llm_model: Optional[str] = None) -> Tuple[pd.DataFrame, HeaderAnalysis, Dict[str, Dict]]:
        """
        è½¬æ¢ä¸ºDataFrame
        
        å‚æ•°:
            analysis: é¢„å…ˆçš„åˆ†æç»“æœï¼Œå¦‚æœä¸ºNoneåˆ™è‡ªåŠ¨åˆ†æ
            use_llm_validate: æ˜¯å¦ä½¿ç”¨LLMéªŒè¯è§„åˆ™åˆ†æç»“æœ
            llm_api_key: LLM APIå¯†é’¥ï¼ˆå¯é€‰ï¼‰
            llm_base_url: LLM APIåœ°å€ï¼ˆå¯é€‰ï¼‰
            llm_model: LLMæ¨¡å‹åç§°ï¼ˆå¯é€‰ï¼‰
        
        è¿”å›:
            (DataFrame, åˆ†æç»“æœ, åˆ—ç»“æ„å…ƒæ•°æ®)
        """
        if analysis is None:
            # å…ˆè¿›è¡Œè§„åˆ™åˆ†æï¼ˆåªåšè¡Œæ£€æµ‹ï¼Œä¸åšåˆ—æ£€æµ‹ï¼‰
            analysis = self.analyze_with_rules()
            
            # å¦‚æœå¯ç”¨LLMéªŒè¯ï¼Œç”¨LLMéªŒè¯è§„åˆ™åˆ†æç»“æœ
            # ä¼˜å…ˆä½¿ç”¨ä¼ å…¥çš„é…ç½®ï¼Œå¦åˆ™ä½¿ç”¨å…¨å±€é…ç½®
            api_key = llm_api_key if llm_api_key is not None else EXCEL_LLM_API_KEY
            if use_llm_validate and api_key:
                analysis = self.validate_with_llm(analysis, llm_api_key, llm_base_url, llm_model)
        
        # LLMéªŒè¯å®Œæˆåï¼Œè¿›è¡Œåˆ—æ£€æµ‹ï¼ˆä½¿ç”¨æœ€ç»ˆçš„è¡¨å¤´è¡Œæ•°å’Œæ•°æ®èµ·å§‹è¡Œï¼‰
        if analysis.valid_cols is None:
            logger.info("ğŸ” LLMéªŒè¯å®Œæˆï¼Œå¼€å§‹è¿›è¡Œåˆ—æ£€æµ‹...")
            valid_cols = self._detect_valid_columns(
                analysis.skip_rows, 
                analysis.header_rows, 
                analysis.data_start_row
            )
            # æ›´æ–°åˆ†æç»“æœï¼Œæ·»åŠ åˆ—æ£€æµ‹ç»“æœ
            analysis.valid_cols = valid_cols
            logger.info("âœ… åˆ—æ£€æµ‹å®Œæˆ")
        
        headers, column_metadata = self.extract_headers(analysis)
        
        # ç¡®å®šè¦è¯»å–çš„åˆ—ï¼ˆå¦‚æœæŒ‡å®šäº†æœ‰æ•ˆåˆ—ï¼Œåªè¯»å–æœ‰æ•ˆåˆ—ï¼‰
        cols_to_read = analysis.valid_cols if analysis.valid_cols is not None else list(range(1, self.ws.max_column + 1))
        
        logger.info(f"ğŸ“Š è¯»å–æ•°æ®: ä» {len(cols_to_read)} åˆ—è¯»å–æ•°æ®")
        
        # è¯»å–æ•°æ®
        data = []
        for row in range(analysis.data_start_row, self.ws.max_row + 1):
            row_data = []
            for col in cols_to_read:
                row_data.append(self.ws.cell(row, col).value)
            if any(v is not None for v in row_data):
                data.append(row_data)
        
        df = pd.DataFrame(data, columns=headers)
        logger.info(f"âœ… DataFrame åˆ›å»ºå®Œæˆ: {len(df)} è¡Œ x {len(df.columns)} åˆ—")
        return df, analysis, column_metadata
    
    def close(self):
        """å…³é—­å·¥ä½œç°¿"""
        try:
            self.wb.close()
        except Exception:
            pass


def process_excel_file(
    filepath: str,
    output_dir: str,
    sheet_name: str = None,
    use_llm_validate: bool = False,
    output_filename: str = None,
    llm_api_key: Optional[str] = None,
    llm_base_url: Optional[str] = None,
    llm_model: Optional[str] = None
) -> ExcelProcessResult:
    """
    å¤„ç†Excelæ–‡ä»¶çš„ä¸»å‡½æ•°
    
    å‚æ•°:
        filepath: Excelæ–‡ä»¶è·¯å¾„
        output_dir: è¾“å‡ºç›®å½•
        sheet_name: å·¥ä½œè¡¨åç§°
        use_llm_validate: æ˜¯å¦ä½¿ç”¨LLMéªŒè¯è§„åˆ™åˆ†æç»“æœ
        output_filename: è¾“å‡ºæ–‡ä»¶åï¼ˆä¸å«æ‰©å±•åï¼‰
        llm_api_key: LLM APIå¯†é’¥ï¼ˆå¯é€‰ï¼‰
        llm_base_url: LLM APIåœ°å€ï¼ˆå¯é€‰ï¼‰
        llm_model: LLMæ¨¡å‹åç§°ï¼ˆå¯é€‰ï¼‰
    
    è¿”å›:
        ExcelProcessResult
    """
    try:
        # ç¡®ä¿è¾“å‡ºç›®å½•å­˜åœ¨
        os.makedirs(output_dir, exist_ok=True)
        
        # å¤„ç†Excel
        processor = SmartHeaderProcessor(filepath, sheet_name)
        df, analysis, column_metadata = processor.to_dataframe(
            use_llm_validate=use_llm_validate,
            llm_api_key=llm_api_key,
            llm_base_url=llm_base_url,
            llm_model=llm_model
        )
        processor.close()
        
        # ç”Ÿæˆè¾“å‡ºæ–‡ä»¶å
        if not output_filename:
            base_name = Path(filepath).stem
            output_filename = f"{base_name}_processed"
        
        # ä¿å­˜CSV
        csv_path = os.path.join(output_dir, f"{output_filename}.csv")
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        
        # ä¿å­˜å…ƒæ•°æ®
        metadata = {
            "header_analysis": analysis.to_dict(),
            "column_metadata": column_metadata,
            "column_names": list(df.columns),
            "row_count": len(df),
            "original_file": os.path.basename(filepath)
        }
        metadata_path = os.path.join(output_dir, f"{output_filename}_metadata.json")
        with open(metadata_path, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)
        
        return ExcelProcessResult(
            success=True,
            header_analysis=analysis,
            processed_file_path=csv_path,
            metadata_file_path=metadata_path,
            column_names=list(df.columns),
            column_metadata=column_metadata,
            row_count=len(df),
            error_message=None
        )
        
    except Exception as e:
        import traceback
        error_msg = f"{str(e)}\n{traceback.format_exc()}"
        return ExcelProcessResult(
            success=False,
            header_analysis=None,
            processed_file_path=None,
            metadata_file_path=None,
            column_names=[],
            column_metadata={},
            row_count=0,
            error_message=error_msg
        )


def get_sheet_names(filepath: str) -> List[str]:
    """è·å–Excelæ–‡ä»¶çš„æ‰€æœ‰å·¥ä½œè¡¨åç§°"""
    try:
        wb = load_workbook(filepath, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception as e:
        return []


def generate_analysis_prompt(
    process_result: ExcelProcessResult,
    custom_prompt: str = None,
    include_metadata: bool = True
) -> str:
    """
    æ ¹æ®Excelå¤„ç†ç»“æœç”Ÿæˆæ•°æ®åˆ†ææç¤ºè¯
    
    å‚æ•°:
        process_result: Excelå¤„ç†ç»“æœ
        custom_prompt: è‡ªå®šä¹‰åˆ†ææç¤ºè¯
        include_metadata: æ˜¯å¦åŒ…å«åˆ—ç»“æ„å…ƒæ•°æ®
    
    è¿”å›:
        æ ¼å¼åŒ–çš„æç¤ºè¯
    """
    if not process_result.success:
        return ""
    
    # åŸºç¡€ä¿¡æ¯
    prompt_parts = []
    
    if custom_prompt:
        prompt_parts.append(custom_prompt)
    else:
        prompt_parts.append("è¯·å¯¹ä¸Šä¼ çš„æ•°æ®è¿›è¡Œå…¨é¢åˆ†æï¼Œç”Ÿæˆæ•°æ®åˆ†ææŠ¥å‘Šã€‚")
    
    # æ·»åŠ æ•°æ®æ¦‚å†µ
    prompt_parts.append(f"\n\n## æ•°æ®æ¦‚å†µ")
    prompt_parts.append(f"- æ•°æ®è¡Œæ•°: {process_result.row_count}")
    prompt_parts.append(f"- åˆ—æ•°: {len(process_result.column_names)}")
    prompt_parts.append(f"- åˆ—å: {', '.join(process_result.column_names[:20])}")
    if len(process_result.column_names) > 20:
        prompt_parts.append(f"  ... ç­‰å…± {len(process_result.column_names)} åˆ—")
    
    # æ·»åŠ è¡¨å¤´åˆ†æä¿¡æ¯
    if process_result.header_analysis:
        ha = process_result.header_analysis
        prompt_parts.append(f"\n## è¡¨å¤´ç»“æ„")
        prompt_parts.append(f"- è¡¨å¤´ç±»å‹: {ha.header_type}")
        if ha.header_type == 'multi':
            prompt_parts.append(f"- è¡¨å¤´å±‚çº§: {ha.header_rows}å±‚")
    
    # æ·»åŠ åˆ—ç»“æ„å…ƒæ•°æ®ï¼ˆå¸®åŠ©AIç†è§£åˆ—ä¹‹é—´çš„å…³ç³»ï¼‰
    if include_metadata and process_result.column_metadata:
        # æ£€æŸ¥æ˜¯å¦æœ‰å¤šçº§ç»“æ„
        has_multi_level = any(
            len(meta) > 1 
            for meta in process_result.column_metadata.values()
        )
        
        if has_multi_level:
            prompt_parts.append(f"\n## åˆ—å±‚çº§ç»“æ„ï¼ˆå¸®åŠ©ç†è§£åˆ—ä¹‹é—´çš„åˆ†ç»„å…³ç³»ï¼‰")
            # æŒ‰level1åˆ†ç»„å±•ç¤º
            groups = defaultdict(list)
            for col_name, meta in process_result.column_metadata.items():
                level1 = meta.get('level1', col_name)
                groups[level1].append(col_name)
            
            for group, cols in groups.items():
                if len(cols) > 1:
                    prompt_parts.append(f"- {group}: {', '.join(cols)}")
    
    return '\n'.join(prompt_parts)

