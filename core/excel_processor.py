"""
Excel智能处理模块
支持：
1. 自动跳过无效行（注释、标题等）
2. 单表头/多表头自动识别
3. 可选调用LLM进行智能分析
4. 合并单元格处理
5. 列结构元数据生成
"""

import pandas as pd
import json
import re
import os
import sys
import requests
import logging
import tempfile
import shutil
import time
import zipfile
from openpyxl import load_workbook
from typing import Tuple, List, Dict, Optional, Any
from collections import defaultdict
from dataclasses import dataclass, asdict, field
from pathlib import Path

# 配置日志
logger = logging.getLogger(__name__)

# 导入配置（避免循环导入，使用延迟导入）

from .config import EXCEL_LLM_API_KEY, EXCEL_LLM_BASE_URL, EXCEL_LLM_MODEL



@dataclass
class HeaderAnalysis:
    """表头分析结果"""
    skip_rows: int          # 需要跳过的无效行数
    header_rows: int        # 表头占用的行数
    header_type: str        # 'single' 或 'multi'
    data_start_row: int     # 数据开始行（1-indexed）
    confidence: str         # 置信度: high/medium/low
    reason: str             # 分析原因说明
    start_col: int = 1      # 数据起始列（1-indexed），第一个表头行中第一个非空表头开始的列
    valid_cols: Optional[List[int]] = None  # 有效列的索引列表（1-indexed），None表示所有列都有效
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        result = asdict(self)
        if result.get('valid_cols') is None:
            result['valid_cols'] = None
        return result


@dataclass
class ExcelProcessResult:
    """Excel处理结果"""
    success: bool
    header_analysis: Optional[HeaderAnalysis]
    processed_file_path: Optional[str]      # 处理后的CSV文件路径
    metadata_file_path: Optional[str]       # 元数据JSON文件路径
    column_names: List[str]                 # 列名列表
    column_metadata: Dict[str, Dict]        # 列结构元数据
    row_count: int                          # 数据行数
    error_message: Optional[str]            # 错误信息
    llm_analysis_response: Optional[str] = None  # LLM分析原始响应（用于调试）
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        return {
            "success": self.success,
            "header_analysis": self.header_analysis.to_dict() if self.header_analysis else None,
            "processed_file_path": self.processed_file_path,
            "metadata_file_path": self.metadata_file_path,
            "column_names": self.column_names,
            "column_metadata": self.column_metadata,
            "row_count": self.row_count,
            "error_message": self.error_message
        }


class SmartHeaderProcessor:
    """智能表头处理器"""
    
    def __init__(self, filepath: str, sheet_name: str = None, load_timeout: int = 60, read_timeout: int = 10, debug_print_header_analysis: bool = False, max_file_size_mb: Optional[int] = None, max_rows: Optional[int] = None):
        """
        初始化智能表头处理器
        
        参数:
            filepath: Excel文件路径
            sheet_name: 工作表名称（可选）
            load_timeout: 加载Excel文件的超时时间（秒），默认60秒
            read_timeout: 读取Excel数据的超时时间（秒），默认10秒
            debug_print_header_analysis: 是否流式打印原始数据（用于调试），默认False
            max_file_size_mb: 最大文件大小（MB），如果为None则使用默认值
        """
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.file_ext = Path(filepath).suffix.lower()
        self._temp_xlsx_path = None  # 用于存储临时转换的 .xlsx 文件路径
        self.read_timeout = read_timeout  # 读取数据的超时时间
        self.debug_print_header_analysis = debug_print_header_analysis  # 是否流式打印原始数据
        
        # 文件预检查（在加载之前）
        logger.info(f"🔍 [DEBUG] SmartHeaderProcessor.__init__: 开始文件预检查")
        try:
            # 基础检查（文件存在、大小、可读）
            _validate_excel_file_basic(filepath, max_file_size_mb=max_file_size_mb)
            logger.info(f"✅ [DEBUG] SmartHeaderProcessor.__init__: 基础检查通过")
            
            # 如果是 .xlsx 格式，验证ZIP格式
            if self.file_ext == '.xlsx':
                _validate_xlsx_format(filepath, timeout=0.5)
                logger.info(f"✅ [DEBUG] SmartHeaderProcessor.__init__: ZIP格式验证通过")
                
                # Excel结构验证（可选，较慢）
                _validate_excel_structure(filepath, timeout=2.0)
                logger.info(f"✅ [DEBUG] SmartHeaderProcessor.__init__: Excel结构验证通过")
                
                # 行数检查（超过配置的最大行数直接拒绝）
                max_rows_value = max_rows if max_rows is not None else 10000
                _validate_excel_row_count(filepath, sheet_name=sheet_name, max_rows=max_rows_value, timeout=5.0)
                logger.info(f"✅ [DEBUG] SmartHeaderProcessor.__init__: 行数检查通过（限制: {max_rows_value} 行）")
        except Exception as e:
            logger.error(f"❌ [DEBUG] SmartHeaderProcessor.__init__: 文件预检查失败: {e}")
            raise
        
        # 如果是 .xls 格式，先转换为 .xlsx（带超时保护）
        if self.file_ext == '.xls':
            logger.info(f"🔄 检测到 .xls 格式文件，正在转换为 .xlsx...")
            self._temp_xlsx_path = self._convert_xls_to_xlsx(filepath, timeout=load_timeout)
            actual_filepath = self._temp_xlsx_path
            logger.info(f"✅ 转换完成: {self._temp_xlsx_path}")
        else:
            actual_filepath = filepath
        
        # 统一使用 openpyxl 读取（带超时保护）
        # 注意：不使用 read_only 模式，因为需要访问 merged_cells 属性来处理合并单元格
        logger.info(f"⏳ [DEBUG] SmartHeaderProcessor.__init__: 开始加载工作簿，超时: {load_timeout}秒")
        self.wb = self._load_workbook_with_timeout(actual_filepath, timeout=load_timeout)
        logger.info(f"✅ [DEBUG] SmartHeaderProcessor.__init__: 工作簿加载完成")
        # 修复：明确使用第一个工作表，而不是依赖 wb.active（active可能是用户最后查看的工作表）
        logger.info(f"⏳ [DEBUG] SmartHeaderProcessor.__init__: 开始选择工作表")
        if sheet_name:
            self.ws = self.wb[sheet_name]
        else:
            # 明确使用第一个工作表（索引0），确保行为一致
            if not self.wb.sheetnames:
                raise ValueError("Excel文件不包含任何工作表")
            self.ws = self.wb[self.wb.sheetnames[0]]
        logger.info(f"✅ [DEBUG] SmartHeaderProcessor.__init__: 工作表选择完成")
        # 构建合并单元格映射（带超时保护）
        logger.info(f"⏳ [DEBUG] SmartHeaderProcessor.__init__: 开始构建合并单元格映射，超时: {load_timeout}秒")
        self.merged_cells_map = self._build_merged_cells_map_with_timeout(timeout=load_timeout)
        logger.info(f"✅ [DEBUG] SmartHeaderProcessor.__init__: 合并单元格映射构建完成")
    
    def _load_workbook_with_timeout(self, filepath: str, timeout: int = 60):
        """带超时保护的 load_workbook"""
        from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
        
        def _load():
            """在后台线程中加载工作簿"""
            try:
                # 二次验证（快速检查，因为可能已经在 __init__ 中检查过）
                # 只做基础检查，不做ZIP验证（避免重复）
                file_ext = Path(filepath).suffix.lower()
                if file_ext == '.xlsx':
                    # 快速ZIP头验证
                    with open(filepath, 'rb') as f:
                        header = f.read(4)
                        if header != b'PK\x03\x04':
                            raise ValueError(f"不是有效的Excel文件（ZIP格式错误）: {filepath}")
                
                return load_workbook(filepath, data_only=True)
            except Exception as e:
                logger.error(f"加载Excel文件失败: {filepath}, 错误: {e}")
                raise
        
        try:
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(_load)
                try:
                    wb = future.result(timeout=timeout)
                    return wb
                except FutureTimeoutError:
                    logger.error(f"加载Excel文件超时: {filepath} (超时时间: {timeout}秒)")
                    future.cancel()
                    raise TimeoutError(f"加载Excel文件超时（{timeout}秒）: {filepath}")
        except Exception as e:
            if isinstance(e, TimeoutError):
                raise
            logger.error(f"加载Excel文件时发生异常: {filepath}, 错误: {e}")
            raise
    
    def _convert_xls_to_xlsx(self, xls_path: str, timeout: int = 60) -> str:
        """
        将 .xls 文件转换为 .xlsx 格式（带超时保护）
        
        参数:
            xls_path: .xls 文件路径
            timeout: 超时时间（秒），默认60秒
        
        返回:
            临时 .xlsx 文件路径
        """
        from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
        
        def _convert():
            """在后台线程中执行转换"""
            try:
                # 读取所有工作表
                excel_file = pd.ExcelFile(xls_path, engine='xlrd')
                
                # 创建临时文件
                temp_dir = os.path.dirname(xls_path)
                temp_xlsx_path = os.path.join(
                    temp_dir, 
                    f"{Path(xls_path).stem}_converted_{os.getpid()}.xlsx"
                )
                
                # 使用 ExcelWriter 写入所有工作表
                with pd.ExcelWriter(temp_xlsx_path, engine='openpyxl') as writer:
                    for sheet_name in excel_file.sheet_names:
                        df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='xlrd')
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                logger.info(f"✅ .xls 文件已转换为 .xlsx: {temp_xlsx_path}")
                return temp_xlsx_path
            except Exception as e:
                logger.error(f"❌ 转换 .xls 文件失败: {e}")
                raise ValueError(
                    f"无法转换 .xls 文件。请确保已安装 xlrd 库: pip install xlrd。错误: {str(e)}"
                )
        
        try:
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(_convert)
                try:
                    result = future.result(timeout=timeout)
                    return result
                except FutureTimeoutError:
                    logger.error(f"转换 .xls 文件超时: {xls_path} (超时时间: {timeout}秒)")
                    future.cancel()
                    raise TimeoutError(f"转换 .xls 文件超时（{timeout}秒）: {xls_path}")
        except Exception as e:
            if isinstance(e, TimeoutError):
                raise
            logger.error(f"转换 .xls 文件时发生异常: {xls_path}, 错误: {e}")
            raise
    
    def _build_merged_cells_map(self) -> Dict[Tuple[int, int], str]:
        """构建合并单元格映射（内部方法，不带超时）"""
        merged_map = {}
        try:
            for merged_range in self.ws.merged_cells.ranges:
                min_row, min_col = merged_range.min_row, merged_range.min_col
                value = self.ws.cell(min_row, min_col).value
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_map[(row, col)] = value
        except Exception as e:
            # 如果无法获取合并单元格信息，记录警告并返回空字典
            logger.warning(f"⚠️ 构建合并单元格映射时出错: {e}，将使用空映射")
        
        return merged_map
    
    def _build_merged_cells_map_with_timeout(self, timeout: int = 10) -> Dict[Tuple[int, int], str]:
        """带超时保护的构建合并单元格映射"""
        from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
        
        def _build():
            """在后台线程中构建合并单元格映射"""
            try:
                return self._build_merged_cells_map()
            except Exception as e:
                logger.error(f"构建合并单元格映射失败: {e}")
                raise
        
        try:
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(_build)
                try:
                    merged_map = future.result(timeout=timeout)
                    return merged_map
                except FutureTimeoutError:
                    logger.error(f"构建合并单元格映射超时: {self.filepath} (超时时间: {timeout}秒)")
                    future.cancel()
                    # 超时时返回空字典，而不是抛出异常，避免影响后续处理
                    logger.warning(f"⚠️ 构建合并单元格映射超时，将使用空映射")
                    return {}
        except Exception as e:
            # 如果发生其他异常，也返回空字典
            logger.warning(f"⚠️ 构建合并单元格映射时发生异常: {e}，将使用空映射")
            return {}
    
    def get_cell_value(self, row: int, col: int) -> Any:
        """获取单元格值，处理合并单元格"""
        if (row, col) in self.merged_cells_map:
            return self.merged_cells_map[(row, col)]
        return self.ws.cell(row, col).value
    
    def get_preview_data(self, max_rows: int = 15, max_cols: int = 25) -> List[List[Any]]:
        """
        获取预览数据用于分析（简化版）
        
        直接读取原始数据，不做任何处理（包括合并单元格处理）
        带超时保护
        """
        from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
        
        def _read_data():
            """在后台线程中读取数据"""
            actual_max_col = min(self.ws.max_column, max_cols)
            actual_max_row = min(self.ws.max_row, max_rows)
            
            data = []
            for row in range(1, actual_max_row + 1):
                row_data = []
                for col in range(1, actual_max_col + 1):
                    # 直接读取原始值，不做任何处理
                    value = self.ws.cell(row, col).value
                    row_data.append(value)
                data.append(row_data)
            return data
        
        try:
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(_read_data)
                try:
                    data = future.result(timeout=self.read_timeout)
                    
                    # 如果启用调试打印，流式打印原始数据
                    if self.debug_print_header_analysis:
                        print("=" * 80)
                        print("【原始Excel数据 - 流式打印】（前15行，前25列）")
                        print("=" * 80)
                        sys.stdout.flush()
                        
                        for i, row in enumerate(data, 1):
                            print(f"行{i}: {row}")
                            sys.stdout.flush()
                        
                        print("=" * 80)
                        sys.stdout.flush()
                    
                    return data
                except FutureTimeoutError:
                    logger.error(f"读取Excel数据超时: {self.filepath} (超时时间: {self.read_timeout}秒)")
                    future.cancel()
                    raise TimeoutError(f"读取Excel数据超时（{self.read_timeout}秒）: {self.filepath}")
        except Exception as e:
            if isinstance(e, TimeoutError):
                raise
            logger.error(f"读取Excel数据时发生异常: {self.filepath}, 错误: {e}")
            raise
    
    
    def analyze_with_llm(self, 
                         llm_api_key: Optional[str] = None,
                         llm_base_url: Optional[str] = None,
                         llm_model: Optional[str] = None,
                         timeout: Optional[int] = None,
                         thinking_callback: Optional[callable] = None) -> Tuple[HeaderAnalysis, str]:
        """
        使用LLM直接分析Excel表格的行和列结构（简化版）
        
        参数:
            llm_api_key: LLM API密钥（可选）
            llm_base_url: LLM API地址（可选）
            llm_model: LLM模型名称（可选）
            timeout: 超时时间（秒），默认90秒
            thinking_callback: 用于流式输出 thinking 内容的回调函数（可选）
        
        返回:
            (分析结果, LLM原始响应)（如果LLM调用失败，抛出异常）
        """
        # 直接读取前15行、25列的原始数据，不做任何处理
        preview_data = self.get_preview_data(max_rows=15, max_cols=25)
        max_col = self.ws.max_column
        
        # 构建简化的分析提示词
        prompt = self._build_llm_analysis_prompt(preview_data, max_col)
        
        # 调用LLM（使用传入的配置或从全局配置读取）
        result = self._call_llm(prompt, llm_api_key, llm_base_url, llm_model, timeout=timeout, thinking_callback=thinking_callback)
        
        if not result:
            raise ValueError("LLM分析失败：无法获取LLM响应，请检查API配置")
        
        # 解析LLM分析结果
        analysis = self._parse_llm_analysis_response(result)
        
        return analysis, result
    
    def validate_with_llm(self, rule_analysis: HeaderAnalysis, 
                         llm_api_key: Optional[str] = None,
                         llm_base_url: Optional[str] = None,
                         llm_model: Optional[str] = None,
                         timeout: Optional[int] = None) -> HeaderAnalysis:
        """
        使用LLM验证规则分析的结果（已废弃，保留用于兼容性）
        
        参数:
            rule_analysis: 规则分析的结果
            llm_api_key: LLM API密钥（可选）
            llm_base_url: LLM API地址（可选）
            llm_model: LLM模型名称（可选）
            timeout: 超时时间（秒），默认30秒
        
        返回:
            验证后的分析结果（如果LLM验证失败，返回原规则分析结果）
        """
        preview_data = self.get_preview_data()
        
        # 构建验证提示词
        prompt = self._build_validation_prompt(preview_data, rule_analysis)
        
        # 调用LLM（使用传入的配置或从全局配置读取）
        result = self._call_llm(prompt, llm_api_key, llm_base_url, llm_model, timeout=timeout)
        
        # 解析LLM验证结果
        validated = self._parse_validation_response(result, rule_analysis)
        
        return validated
    
    def _build_llm_analysis_prompt(self, preview_data: List[List], 
                                   max_col: int) -> str:
        """
        构建LLM分析提示词（简化版）
        
        直接读取原始数据，不做任何处理，让LLM直接识别
        """
        # 格式化预览数据为简单的表格形式
        num_cols = len(preview_data[0]) if preview_data else 0
        num_rows = len(preview_data)
        
        # 构建简单的表格字符串
        table_str = "【Excel原始数据】（前15行，前25列）\n\n"
        table_str += "行号 | " + " | ".join([f"列{i+1}" for i in range(num_cols)]) + "\n"
        table_str += "-" * (8 + num_cols * 15) + "\n"
        
        for i, row in enumerate(preview_data, 1):
            row_str = " | ".join([str(cell) if cell is not None else "" for cell in row])
            table_str += f"  {i:2d}  | {row_str}\n"
        
        prompt = f"""你是一个Excel表格结构分析专家。请分析以下Excel表格的原始数据，识别表头结构。

{table_str}

【总列数】{max_col}

## 分析任务

请分析表格结构，识别：
1. **无效行（skip_rows）**：表头之前的无效行（如文档标题、说明文字、注释、公司名称、填报说明等）
2. **表头行数（header_rows）**：所有表头行，包括多级表头的所有层级
3. **表头类型（header_type）**：single（单表头）或 multi（多级表头）
4. **数据起始行（data_start_row）**：数据开始的行号，必须等于 skip_rows + header_rows + 1
5. **数据起始列（start_col）**：第一个表头行中第一个非空表头开始的列号

## 识别规则

### 无效行特征：
- 文档标题（如"2024年度报表"）
- 公司名称或部门名称（如"XX公司"、"XX部门"）
- 填报说明（如"填报机构"、"填报日期"、"填报机构/日期"等，任何包含"填报"关键词的行）
- 只有数字没有标签的行（如只有"222"、"111"等数字，没有对应的列名）
- 完全空行或只有少量文本的行

### 表头行特征：
- 包含列名或分类标签（如"销售事业部"、"华东大区"、"线上销售额"等）
- 有明确的层级结构（多级表头）
- 通常不包含大量数值数据

### 数据行特征：
- 包含大量数值数据
- 不再是表头文本或分类标签

## 输出格式

请以JSON格式返回分析结果：

```json
{{
    "skip_rows": <表头之前的无效行数，如果第1行就是表头则填0>,
    "header_rows": <表头占用的总行数>,
    "header_type": "<single或multi>",
    "data_start_row": <数据开始行号（1-indexed），必须等于skip_rows+header_rows+1>,
    "start_col": <数据起始列号（1-indexed）>,
    "valid_cols": null,
    "confidence": "<high/medium/low>",
    "reason": "<详细说明识别过程>"
}}
```

## 注意事项

1. 行号和列号都从1开始计数
2. data_start_row 必须等于 skip_rows + header_rows + 1
3. valid_cols 始终设为 null
4. 只返回JSON，不要其他内容
5. 如果第1行就是表头，则 skip_rows=0
6. 多级表头的所有行都要计入 header_rows

只返回JSON，不要其他内容。"""
        
        return prompt
    
    def _build_validation_prompt(self, preview_data: List[List], 
                                rule_analysis: HeaderAnalysis) -> str:
        """
        构建LLM验证提示词（简化版）
        """
        # 格式化预览数据为简单的表格形式
        num_cols = len(preview_data[0]) if preview_data else 0
        
        table_str = "【Excel原始数据】（前15行，前25列）\n\n"
        table_str += "行号 | " + " | ".join([f"列{i+1}" for i in range(num_cols)]) + "\n"
        table_str += "-" * (8 + num_cols * 15) + "\n"
        
        for i, row in enumerate(preview_data, 1):
            row_str = " | ".join([str(cell) if cell is not None else "" for cell in row])
            table_str += f"  {i:2d}  | {row_str}\n"
        
        prompt = f"""请验证以下Excel表格的表头分析结果是否正确。

{table_str}

【当前分析结果】
- 跳过行数(skip_rows): {rule_analysis.skip_rows}
- 表头行数(header_rows): {rule_analysis.header_rows}
- 表头类型: {rule_analysis.header_type}
- 数据起始行: {rule_analysis.data_start_row} （应该等于 skip_rows + header_rows + 1）
- 分析原因: {rule_analysis.reason}

请验证这个结果是否合理，并以JSON格式返回：
{{
    "is_valid": <true或false>,
    "confidence": "<high/medium/low>",
    "suggestions": {{
        "skip_rows": <建议的跳过行数>,
        "header_rows": <建议的表头行数>,
        "header_type": "<single或multi>",
        "data_start_row": <建议的数据起始行>
    }},
    "reason": "<验证说明>"
}}

## 验证要点

1. skip_rows 只计算表头之前的无效行（如文档标题、注释等），不要把表头行算作skip_rows
2. header_rows 应该包含所有表头行，包括多级表头的所有行
3. data_start_row 必须等于 skip_rows + header_rows + 1

只返回JSON，不要其他内容。"""
        
        return prompt
    
    def _call_llm(self, prompt: str, llm_api_key: Optional[str] = None, 
                  llm_base_url: Optional[str] = None, llm_model: Optional[str] = None,
                  timeout: Optional[int] = None, thinking_callback: Optional[callable] = None) -> str:
        """调用LLM API（支持OpenAI兼容接口）
        
        参数:
            prompt: 提示词
            llm_api_key: LLM API密钥（可选，如果不提供则从配置读取）
            llm_base_url: LLM API地址（可选，如果不提供则从配置读取）
            llm_model: LLM模型名称（可选，如果不提供则从配置读取）
            timeout: 超时时间（秒），默认30秒
            thinking_callback: 用于流式输出 thinking 内容的回调函数（可选）
        """
        # 提供默认回调函数，确保 thinking 内容总是被推送（不检查条件）
        if thinking_callback is None:
            # 默认回调函数：只输出到控制台（不推送到插件）
            def default_callback(content: str):
                pass  # 空回调，不执行任何操作
            thinking_callback = default_callback
        # 优先使用传入的参数，否则从配置读取
        api_key = llm_api_key if llm_api_key is not None else EXCEL_LLM_API_KEY
        base_url = llm_base_url if llm_base_url is not None else EXCEL_LLM_BASE_URL
        model = llm_model if llm_model is not None else EXCEL_LLM_MODEL
        
        logger.info("=" * 60)
        logger.info("🤖 调用 LLM API 进行Excel表格分析")
        logger.info(f"🔗 EXCEL_LLM_BASE_URL: {base_url}")
        logger.info(f"📌 模型: {model}")
        logger.info(f"🔑 API Key: {'已配置' if api_key else '未配置'}")
        logger.info("💭 Thinking 流式输出: 已启用（默认开启）")
        
        if not api_key:
            logger.error("❌ 未配置 LLM API Key，无法进行分析")
            return None
            
        url = base_url
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        # 使用流式调用以支持 thinking 功能（默认启用）
        base_payload = {
            "model": model,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.4,
            "max_tokens": 1000,  # 增加token数量以支持更详细的分析
            "stream": True,  # 流式调用（必须启用以支持 thinking）
        }
        
        # 使用传入的超时时间，默认90秒
        request_timeout = timeout if timeout is not None else 90
        
        logger.info(f"📡 发送 LLM API 请求到: {url} (流式调用)")
        logger.info(f"📝 提示词长度: {len(prompt)} 字符")
        logger.info(f"⏱️ 超时设置: {request_timeout} 秒")
        
        try:
            # 默认启用 thinking 功能（流式输出）
            payload_with_thinking = base_payload.copy()
            payload_with_thinking["enable_thinking"] = True  # 默认启用 thinking
            
            logger.info("💭 已启用 Thinking 功能，将实时流式输出思考过程")
            logger.debug(f"📦 请求 payload (启用 thinking): {json.dumps(payload_with_thinking, ensure_ascii=False, indent=2)}")
            
            response = requests.post(
                url, 
                headers=headers, 
                json=payload_with_thinking, 
                timeout=request_timeout,
                stream=True  # 启用流式响应
            )
            
            # 如果启用 thinking 失败，回退到不使用 thinking
            if response.status_code != 200:
                try:
                    error_json = response.json()
                    if "enable_thinking" in str(error_json).lower():
                        logger.warning("⚠️ API 不支持 enable_thinking 参数，将回退到不使用 thinking")
                        logger.warning("💭 注意：Thinking 流式输出将不可用（API 不支持）")
                        payload_no_thinking = base_payload.copy()
                        logger.debug(f"📦 请求 payload (不使用 thinking): {json.dumps(payload_no_thinking, ensure_ascii=False, indent=2)}")
                        response = requests.post(
                            url, 
                            headers=headers, 
                            json=payload_no_thinking, 
                            timeout=request_timeout,
                            stream=True
                        )
                except:
                    pass
            
            # 如果请求失败，输出详细的错误信息
            if response.status_code != 200:
                error_detail = ""
                try:
                    # 对于流式响应，尝试读取错误信息
                    error_text = ""
                    for line in response.iter_lines():
                        if line:
                            line_str = line.decode('utf-8')
                            if line_str.startswith('data: '):
                                line_str = line_str[6:]
                            try:
                                error_json = json.loads(line_str)
                                error_detail = json.dumps(error_json, ensure_ascii=False, indent=2)
                                break
                            except:
                                error_text += line_str + "\n"
                    if not error_detail:
                        error_detail = error_text or response.text
                except:
                    try:
                        error_detail = response.text
                    except:
                        error_detail = f"无法读取错误详情 (状态码: {response.status_code})"
                
                logger.error(f"❌ LLM API 调用失败 (状态码: {response.status_code})")
                logger.error(f"📋 错误详情:\n{error_detail}")
                logger.error(f"🔗 请求 URL: {url}")
                logger.error(f"📦 请求 payload: {json.dumps(base_payload, ensure_ascii=False, indent=2)}")
                return None
            
            # 处理流式响应
            full_content = ""
            full_thinking = ""  # 保存完整的 thinking 内容
            thinking_started = False  # 标记是否已经开始输出 thinking
            
            logger.info("=" * 60)
            logger.info("🧠 开始接收 LLM 流式响应（包含 thinking 过程）")
            logger.info("💭 Thinking 流式输出已启用，将实时显示思考过程")
            logger.info("=" * 60)
            
            # 用于调试：记录第一个 chunk 的完整结构
            first_chunk_logged = False
            
            for line in response.iter_lines():
                if line:
                    line_str = line.decode('utf-8')
                    # 跳过 SSE 格式的前缀 "data: "
                    if line_str.startswith('data: '):
                        line_str = line_str[6:]
                    
                    # 检查是否是结束标记
                    if line_str.strip() == '[DONE]':
                        break
                    
                    # 解析 JSON
                    try:
                        chunk_data = json.loads(line_str)
                        
                        # 调试：输出第一个 chunk 的完整结构（帮助了解 API 响应格式）
                        if not first_chunk_logged:
                            logger.info("=" * 60)
                            logger.info("🔍 第一个 Chunk 完整结构（用于调试）:")
                            logger.info("=" * 60)
                            logger.info(json.dumps(chunk_data, ensure_ascii=False, indent=2))
                            logger.info("=" * 60)
                            first_chunk_logged = True
                        
                        if 'choices' in chunk_data and len(chunk_data['choices']) > 0:
                            choice = chunk_data['choices'][0]
                            delta = choice.get('delta', {})
                            finish_reason = choice.get('finish_reason')
                            
                            # 检测是否有 thinking 相关字段（即使内容为空也要检测）
                            has_thinking_field = False
                            thinking_content = None
                            
                            # 方式1: delta.reasoning_content（Qwen 等模型使用）
                            if 'reasoning_content' in delta:
                                has_thinking_field = True
                                thinking_content = delta.get('reasoning_content', '')
                            
                            # 方式2: delta.thinking（最常见）
                            elif 'thinking' in delta:
                                has_thinking_field = True
                                thinking_content = delta.get('thinking', '')
                            
                            # 方式3: delta.reasoning（某些 API 使用）
                            elif 'reasoning' in delta:
                                has_thinking_field = True
                                thinking_content = delta.get('reasoning', '')
                            
                            # 方式4: choice 中直接有 thinking 字段
                            elif 'thinking' in choice:
                                has_thinking_field = True
                                thinking_content = choice.get('thinking', '')
                            
                            # 方式5: finish_reason 为 thinking 时，整个 delta 可能是 thinking
                            elif finish_reason == 'thinking':
                                has_thinking_field = True
                                # 如果 finish_reason 是 thinking，尝试从 delta 中提取
                                if delta:
                                    # 尝试获取所有非 content 的字段作为 thinking
                                    thinking_dict = {k: v for k, v in delta.items() if k != 'content' and k != 'role'}
                                    if thinking_dict:
                                        # 优先使用 reasoning_content
                                        if 'reasoning_content' in thinking_dict:
                                            thinking_content = thinking_dict['reasoning_content']
                                        else:
                                            thinking_content = json.dumps(thinking_dict, ensure_ascii=False)
                                    else:
                                        thinking_content = str(delta)
                            
                            # 方式6: 检查整个 chunk_data 中是否有 thinking 字段
                            elif 'thinking' in chunk_data:
                                has_thinking_field = True
                                thinking_content = chunk_data.get('thinking', '')
                            
                            # 方式7: 检查整个 chunk_data 中是否有 reasoning_content 字段
                            elif 'reasoning_content' in chunk_data:
                                has_thinking_field = True
                                thinking_content = chunk_data.get('reasoning_content', '')
                            
                            # 如果检测到 thinking 字段（即使内容为空），标记为已开始
                            if has_thinking_field and not thinking_started:
                                thinking_prefix = "💭 [Thinking] "
                                logger.info("💭 开始输出 Thinking 过程...")
                                # 总是调用回调函数（不检查条件，确保总是推送）
                                thinking_callback(thinking_prefix)
                                # 初始化日志计数器
                                if not hasattr(self, '_thinking_log_count'):
                                    self._thinking_log_count = 0
                                self._thinking_log_count += 1
                                logger.info(f"💭 [DEBUG] 已调用 thinking_callback 推送前缀 #{self._thinking_log_count}: '{thinking_prefix}'")
                                thinking_started = True
                            
                            # 实时输出 thinking 内容（立即输出，不积累）
                            # 注意：只要检测到 thinking 内容（包括空字符串），就立即推送
                            if thinking_content is not None:
                                # 确保 thinking_content 是字符串
                                if not isinstance(thinking_content, str):
                                    thinking_content = str(thinking_content)
                                
                                # 累积 thinking 内容（用于后续处理）
                                full_thinking += thinking_content
                                # 总是调用回调函数（不检查条件，确保总是推送，即使内容为空）
                                thinking_callback(thinking_content)
                                
                                # 减少日志频率：每30个chunk记录一次，或内容长度 > 50 时记录
                                if not hasattr(self, '_thinking_log_count'):
                                    self._thinking_log_count = 0
                                self._thinking_log_count += 1
                                if self._thinking_log_count % 30 == 1 or len(thinking_content) > 50:
                                    logger.info(f"💭 [DEBUG] 已调用 thinking_callback 推送内容 #{self._thinking_log_count}: {len(thinking_content)} 字符, 内容预览: '{thinking_content[:100] if len(thinking_content) > 100 else thinking_content}'")
                            
                            # 提取普通 content 内容
                            content = delta.get('content', '')
                            if content:
                                full_content += content
                            
                            # 调试：输出 chunk 结构（仅在 debug 模式下）
                            if logger.isEnabledFor(logging.DEBUG):
                                logger.debug(f"📦 Chunk: finish_reason={finish_reason}, delta_keys={list(delta.keys())}, has_thinking={'thinking' in delta or 'thinking' in choice}, has_content=bool(content)")
                            
                    except json.JSONDecodeError:
                        # 忽略无法解析的行（可能是空行或其他格式）
                        continue
            
            # Thinking 流式输出完成
            if thinking_started:
                logger.info("💭 Thinking 流式输出完成")
            elif first_chunk_logged and not full_thinking:
                # 如果收到了 chunk 但没有 thinking 内容，可能是 API 不支持或模型未生成 thinking
                logger.info("💭 注意：已接收响应但未检测到 Thinking 内容（可能 API 不支持或模型未生成 thinking）")
            
            # 输出完整的 thinking 过程（如果有）
            if full_thinking:
                logger.info("=" * 60)
                logger.info("🧠 LLM Thinking 过程（完整）:")
                logger.info("=" * 60)
                logger.info(full_thinking)
                logger.info("=" * 60)
            
            if not full_content:
                logger.warning("⚠️ LLM 流式响应为空")
                return None
            
            logger.info("✅ LLM API 调用成功")
            logger.info("=" * 60)
            logger.info("📝 LLM 响应内容:")
            logger.info("=" * 60)
            logger.info(full_content)
            logger.info("=" * 60)
            
            return full_content
        except requests.exceptions.RequestException as e:
            logger.error(f"❌ LLM调用失败 (网络错误): {e}")
            if hasattr(e, 'response') and e.response is not None:
                try:
                    error_json = e.response.json()
                    logger.error(f"📋 API 错误响应: {json.dumps(error_json, ensure_ascii=False, indent=2)}")
                except:
                    logger.error(f"📋 API 错误响应 (文本): {e.response.text}")
            logger.debug("异常详情:", exc_info=True)
            return None
        except Exception as e:
            logger.error(f"❌ LLM调用失败: {e}")
            logger.debug("异常详情:", exc_info=True)
            return None
    
    def _parse_llm_analysis_response(self, response: str) -> HeaderAnalysis:
        """解析LLM分析结果（包含行和列信息）"""
        if not response:
            raise ValueError("LLM响应为空")
        
        try:
            # 提取JSON部分（支持嵌套JSON）
            # 先尝试找到第一个 { 到最后一个 } 之间的内容
            start_idx = response.find('{')
            end_idx = response.rfind('}')
            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                json_str = response[start_idx:end_idx + 1]
                data = json.loads(json_str)
            else:
                # 如果找不到完整的JSON，尝试用正则匹配
                json_match = re.search(r'\{.*\}', response, re.DOTALL)
                if not json_match:
                    raise ValueError("未找到JSON格式的响应")
                data = json.loads(json_match.group())
            
            # 解析有效列
            valid_cols = data.get('valid_cols')
            if valid_cols is None:
                # 如果为null，表示所有列都有效
                valid_cols = None
            elif isinstance(valid_cols, list):
                # 确保是整数列表
                valid_cols = [int(col) for col in valid_cols if isinstance(col, (int, str)) and str(col).isdigit()]
                # 如果列表为空或包含所有列，设为None
                max_col = self.ws.max_column
                if not valid_cols or set(valid_cols) == set(range(1, max_col + 1)):
                    valid_cols = None
            else:
                valid_cols = None
            
            # 解析起始列（默认为1）
            start_col = int(data.get('start_col', 1))
            if start_col < 1:
                start_col = 1
            
            # 构建HeaderAnalysis对象
            analysis = HeaderAnalysis(
                skip_rows=int(data.get('skip_rows', 0)),
                header_rows=int(data.get('header_rows', 1)),
                header_type=data.get('header_type', 'single'),
                data_start_row=int(data.get('data_start_row', 1)),
                start_col=start_col,
                confidence=data.get('confidence', 'medium'),
                reason=f"LLM分析: {data.get('reason', '')}",
                valid_cols=valid_cols
            )
            
            logger.info(f"✅ LLM分析完成:")
            logger.info(f"  - 跳过行数: {analysis.skip_rows}")
            logger.info(f"  - 表头行数: {analysis.header_rows}")
            logger.info(f"  - 表头类型: {analysis.header_type}")
            logger.info(f"  - 数据起始行: {analysis.data_start_row}")
            logger.info(f"  - 数据起始列: {analysis.start_col}")
            logger.info(f"  - 置信度: {analysis.confidence}")
            
            return analysis
        except (json.JSONDecodeError, KeyError, ValueError) as e:
            logger.error(f"❌ 解析LLM分析响应失败: {e}")
            logger.error(f"📋 响应内容: {response[:500]}")
            raise ValueError(f"解析LLM分析响应失败: {e}")
    
    def _parse_validation_response(self, response: str, rule_analysis: HeaderAnalysis) -> HeaderAnalysis:
        """解析LLM验证结果（已废弃，保留用于兼容性）"""
        if not response:
            # LLM调用失败，返回原规则分析结果
            return rule_analysis
        
        try:
            # 提取JSON部分（支持嵌套JSON）
            # 先尝试找到第一个 { 到最后一个 } 之间的内容
            start_idx = response.find('{')
            end_idx = response.rfind('}')
            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                json_str = response[start_idx:end_idx + 1]
                data = json.loads(json_str)
            else:
                # 如果找不到完整的JSON，尝试用正则匹配
                json_match = re.search(r'\{.*\}', response, re.DOTALL)
                if not json_match:
                    raise ValueError("未找到JSON格式的响应")
                data = json.loads(json_match.group())
            
            is_valid = data.get('is_valid', True)
            suggestions = data.get('suggestions', {})
            
            if is_valid:
                # LLM认为规则分析结果合理，保持原结果但更新置信度和原因
                return HeaderAnalysis(
                    skip_rows=rule_analysis.skip_rows,
                    header_rows=rule_analysis.header_rows,
                    header_type=rule_analysis.header_type,
                    data_start_row=rule_analysis.data_start_row,
                    start_col=rule_analysis.start_col,  # 保持原有的起始列
                    confidence=data.get('confidence', 'high'),  # LLM验证通过，置信度提升
                    reason=f"规则分析+LLM验证: {data.get('reason', '验证通过')}",
                    valid_cols=rule_analysis.valid_cols  # 保持原有的列过滤结果
                )
            else:
                # LLM认为不合理，使用LLM的建议
                # 注意：LLM可能建议修改表头行数，但列过滤结果仍然保留
                return HeaderAnalysis(
                    skip_rows=suggestions.get('skip_rows', rule_analysis.skip_rows),
                    header_rows=suggestions.get('header_rows', rule_analysis.header_rows),
                    header_type=suggestions.get('header_type', rule_analysis.header_type),
                    data_start_row=suggestions.get('data_start_row', rule_analysis.data_start_row),
                    start_col=suggestions.get('start_col', rule_analysis.start_col),  # 保持或使用建议的起始列
                    confidence=data.get('confidence', 'medium'),
                    reason=f"规则分析+LLM修正: {data.get('reason', 'LLM建议修正')}",
                    valid_cols=rule_analysis.valid_cols  # 保持原有的列过滤结果
                )
        except (json.JSONDecodeError, KeyError, ValueError) as e:
            print(f"解析LLM验证响应失败: {e}，使用原规则分析结果")
        
        # 解析失败，返回原规则分析结果
        return rule_analysis
    
    def _detect_valid_columns(self, skip_rows: int, header_rows: int, data_start_row: int) -> List[int]:
        """
        检测有效列（过滤无效列）
        
        无效列的判断标准：
        1. 表头区域完全为空
        2. 数据区域完全为空或没有数值数据
        
        返回: 有效列的索引列表（1-indexed）
        """
        max_col = self.ws.max_column
        header_start = skip_rows + 1
        header_end = skip_rows + header_rows
        valid_cols = []
        
        logger.info("🔍 开始检测无效列...")
        
        for col in range(1, max_col + 1):
            # 检查表头区域是否有内容
            has_header = False
            for row in range(header_start, header_end + 1):
                value = self.get_cell_value(row, col)
                if value is not None and str(value).strip():
                    has_header = True
                    break
            
            # 检查数据区域是否有数值数据
            has_data = False
            numeric_count = 0
            total_count = 0
            for row in range(data_start_row, min(data_start_row + 10, self.ws.max_row + 1)):
                value = self.ws.cell(row, col).value
                if value is not None:
                    total_count += 1
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        numeric_count += 1
                        has_data = True
            
            # 如果表头有内容或数据区域有数值，则认为是有效列
            if has_header or has_data:
                valid_cols.append(col)
                logger.debug(f"✅ 列 {col}: 有效 (表头: {has_header}, 数据: {has_data}, 数值: {numeric_count}/{total_count})")
            else:
                logger.info(f"❌ 列 {col}: 无效 (表头为空且数据为空)")
        
        logger.info(f"📊 列过滤结果: 总列数 {max_col}, 有效列数 {len(valid_cols)}, 无效列数 {max_col - len(valid_cols)}")
        
        # 如果所有列都有效，返回None（表示不需要过滤）
        if len(valid_cols) == max_col:
            return None
        
        return valid_cols
    
    def extract_headers(self, analysis: HeaderAnalysis) -> Tuple[List[str], Dict[str, Dict]]:
        """
        根据分析结果提取表头
        返回: (列名列表, 列结构元数据)
        """
        max_col = self.ws.max_column
        header_start = analysis.skip_rows + 1
        header_end = analysis.skip_rows + analysis.header_rows
        
        # 确定要处理的列：从 start_col 开始，如果指定了有效列，则取交集
        all_cols = list(range(analysis.start_col, max_col + 1))
        if analysis.valid_cols is not None:
            # 取交集：从 start_col 开始，且在 valid_cols 中的列
            cols_to_process = [col for col in all_cols if col in analysis.valid_cols]
        else:
            cols_to_process = all_cols
        
        logger.info(f"📋 提取表头: 处理 {len(cols_to_process)} 列")
        
        # 调试：流式打印原始多级表头（不做任何美化）
        print("=" * 80)
        print("【原始多级表头 - 流式打印】")
        print(f"表头行范围: 第 {header_start} 行到第 {header_end} 行")
        print(f"处理列范围: 第 {analysis.start_col} 列到第 {max_col} 列")
        print("=" * 80)
        sys.stdout.flush()
        
        for row in range(header_start, header_end + 1):
            row_values = []
            for col in cols_to_process:
                value = self.get_cell_value(row, col)
                row_values.append(value)
            # 直接打印，不做任何美化
            print(f"行{row}: {row_values}")
            sys.stdout.flush()
        
        print("=" * 80)
        sys.stdout.flush()
        
        column_metadata = {}
        
        if analysis.header_type == 'single':
            # 单表头
            headers = []
            for col in cols_to_process:
                value = self.get_cell_value(header_start, col)
                col_name = str(value) if value else f'Column_{col}'
                headers.append(col_name)
                column_metadata[col_name] = {"level1": col_name}
            
            headers = self._handle_duplicate_names(headers)
            # 更新元数据的key
            column_metadata = {h: {"level1": h} for h in headers}
            return headers, column_metadata
        
        else:
            # 多表头：展平
            column_headers = []
            original_metadata_list = []  # 保存原始元数据列表，按顺序对应
            
            for col in cols_to_process:
                parts = []
                levels = {}
                for row_idx, row in enumerate(range(header_start, header_end + 1), 1):
                    value = self.get_cell_value(row, col)
                    if value is not None:
                        part = str(value).strip()
                        parts.append(part)
                        levels[f"level{row_idx}"] = part
                
                # 去重连续相同值
                unique_parts = []
                for p in parts:
                    if not unique_parts or p != unique_parts[-1]:
                        unique_parts.append(p)
                
                col_name = '_'.join(unique_parts) if unique_parts else f'Column_{col}'
                column_headers.append(col_name)
                original_metadata_list.append(levels)  # 按顺序保存元数据
            
            # 处理重复列名
            column_headers = self._handle_duplicate_names(column_headers)
            
            # 重新映射元数据：使用索引对应关系
            new_metadata = {}
            for i, header in enumerate(column_headers):
                # 使用索引直接获取对应的元数据
                if i < len(original_metadata_list):
                    new_metadata[header] = original_metadata_list[i]
                else:
                    # 如果索引超出范围，创建默认元数据
                    logger.warning(f"⚠️ 索引超出范围: i={i}, headers长度={len(column_headers)}, metadata长度={len(original_metadata_list)}")
                    new_metadata[header] = {"level1": header}
            
            return column_headers, new_metadata
    
    def _handle_duplicate_names(self, names: List[str]) -> List[str]:
        """处理重复列名"""
        counts = defaultdict(int)
        result = []
        for name in names:
            if counts[name] > 0:
                result.append(f"{name}_{counts[name]}")
            else:
                result.append(name)
            counts[name] += 1
        return result
    
    def to_dataframe(self, analysis: HeaderAnalysis = None, use_llm_validate: bool = False,
                    llm_api_key: Optional[str] = None,
                    llm_base_url: Optional[str] = None,
                    llm_model: Optional[str] = None,
                    preprocessing_timeout: Optional[int] = None,
                    thinking_callback: Optional[callable] = None) -> Tuple[pd.DataFrame, HeaderAnalysis, Dict[str, Dict], Optional[str]]:
        """
        转换为DataFrame
        
        参数:
            analysis: 预先的分析结果，如果为None则使用LLM自动分析
            use_llm_validate: 已废弃，保留用于兼容性
            llm_api_key: LLM API密钥（必填）
            llm_base_url: LLM API地址（可选）
            llm_model: LLM模型名称（可选）
            preprocessing_timeout: 预处理超时时间（秒），默认90秒
        
        返回:
            (DataFrame, 分析结果, 列结构元数据, LLM原始响应)
        """
        llm_response = None
        if analysis is None:
            # 必须使用LLM进行分析（同时分析行和列）
            logger.info("🤖 开始使用LLM分析Excel表格结构（行和列）...")
            
            # 优先使用传入的配置，否则使用全局配置
            api_key = llm_api_key if llm_api_key is not None else EXCEL_LLM_API_KEY
            if not api_key:
                raise ValueError("LLM API密钥未配置，无法进行Excel分析。请配置EXCEL_LLM_API_KEY或传入llm_api_key参数")
            
            # 使用LLM直接分析（包含行和列信息）
            analysis, llm_response = self.analyze_with_llm(
                llm_api_key=llm_api_key,
                llm_base_url=llm_base_url,
                llm_model=llm_model,
                timeout=preprocessing_timeout,
                thinking_callback=thinking_callback
            )
            logger.info("✅ LLM分析完成（已包含行和列信息）")
            # 保存LLM响应到实例变量，以便后续使用
            self._llm_analysis_response = llm_response
        
        headers, column_metadata = self.extract_headers(analysis)
        
        # 确定要读取的列：从 start_col 开始，如果指定了有效列，则取交集
        max_col = self.ws.max_column
        all_cols = list(range(analysis.start_col, max_col + 1))
        if analysis.valid_cols is not None:
            # 取交集：从 start_col 开始，且在 valid_cols 中的列
            cols_to_read = [col for col in all_cols if col in analysis.valid_cols]
        else:
            cols_to_read = all_cols
        
        logger.info(f"📊 读取数据: 从第 {analysis.start_col} 列开始，共 {len(cols_to_read)} 列")
        
        # 读取数据
        data = []
        for row in range(analysis.data_start_row, self.ws.max_row + 1):
            row_data = []
            for col in cols_to_read:
                row_data.append(self.ws.cell(row, col).value)
            if any(v is not None for v in row_data):
                data.append(row_data)
        
        df = pd.DataFrame(data, columns=headers)
        
        # 智能类型转换：尝试将数字字符串转换为数字
        logger.info("🔄 开始智能类型转换...")
        def smart_convert_value(value):
            """智能转换值：尝试将数字字符串转换为数字"""
            if value is None:
                return value
            if isinstance(value, (int, float)):
                return value
            if isinstance(value, str):
                # 去除前后空格
                value = value.strip()
                if not value:  # 空字符串
                    return None
                # 尝试转换为数字
                try:
                    # 尝试整数（支持负数）
                    if value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
                        return int(value)
                    # 尝试浮点数（支持科学计数法）
                    return float(value)
                except (ValueError, AttributeError):
                    # 转换失败，保持原字符串
                    return value
            return value
        
        # 对每列应用智能转换
        for col in df.columns:
            original_type = df[col].dtype
            df[col] = df[col].apply(smart_convert_value)
            new_type = df[col].dtype
            if original_type != new_type:
                logger.debug(f"  列 '{col}': {original_type} → {new_type}")
        
        # 使用 pandas 的 convert_dtypes 进一步优化类型推断
        df = df.convert_dtypes()
        
        logger.info(f"✅ DataFrame 创建完成: {len(df)} 行 x {len(df.columns)} 列")
        logger.info(f"📊 数据类型优化完成")
        return df, analysis, column_metadata, llm_response
    
    def close(self):
        """关闭工作簿并清理临时文件"""
        try:
            self.wb.close()
        except Exception:
            pass
        
        # 删除临时转换的 .xlsx 文件
        if self._temp_xlsx_path and os.path.exists(self._temp_xlsx_path):
            try:
                os.remove(self._temp_xlsx_path)
                logger.debug(f"🗑️ 已删除临时文件: {self._temp_xlsx_path}")
            except Exception as e:
                logger.warning(f"⚠️ 删除临时文件失败: {self._temp_xlsx_path}, 错误: {e}")


def _validate_excel_file_basic(filepath: str, max_file_size_mb: Optional[int] = None) -> None:
    """
    基础文件检查（快速检查）
    
    参数:
        filepath: Excel文件路径
        max_file_size_mb: 最大文件大小（MB），如果为None则使用默认值
    
    异常:
        FileNotFoundError: 文件不存在
        ValueError: 文件为空或过大
        PermissionError: 文件不可读
    """
    from .config import EXCEL_MAX_FILE_SIZE_MB
    
    # 检查文件是否存在
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Excel文件不存在: {filepath}")
    
    # 检查文件是否可读
    if not os.access(filepath, os.R_OK):
        raise PermissionError(f"Excel文件不可读: {filepath}")
    
    # 检查文件大小
    file_size = os.path.getsize(filepath)
    
    # 检查是否为空文件
    if file_size == 0:
        raise ValueError(f"Excel文件为空（0字节）: {filepath}")
    
    # 检查文件大小限制
    max_size_mb = max_file_size_mb if max_file_size_mb is not None else EXCEL_MAX_FILE_SIZE_MB
    max_size_bytes = max_size_mb * 1024 * 1024
    
    if file_size > max_size_bytes:
        file_size_mb = file_size / 1024 / 1024
        raise ValueError(f"Excel文件过大: {file_size_mb:.2f}MB，超过限制（{max_size_mb}MB）: {filepath}")
    
    # 如果文件较大（> 100MB），记录警告
    if file_size > 100 * 1024 * 1024:
        file_size_mb = file_size / 1024 / 1024
        logger.warning(f"⚠️ Excel文件较大: {file_size_mb:.2f}MB，处理可能较慢: {filepath}")


def _validate_xlsx_format(filepath: str, timeout: float = 0.5) -> None:
    """
    验证 .xlsx 文件的ZIP格式（带超时保护）
    
    参数:
        filepath: Excel文件路径
        timeout: 超时时间（秒），默认0.5秒
    
    异常:
        ValueError: ZIP格式错误或文件损坏
    """
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
    
    def _validate():
        """在后台线程中验证ZIP格式"""
        try:
            # 检查ZIP文件头（前4个字节）
            with open(filepath, 'rb') as f:
                header = f.read(4)
                if header != b'PK\x03\x04':
                    raise ValueError(f"不是有效的Excel文件（ZIP格式错误）: {filepath}")
            
            # 尝试打开ZIP文件验证
            try:
                with zipfile.ZipFile(filepath, 'r') as zf:
                    # 尝试读取ZIP文件列表（不实际解压）
                    zf.namelist()
            except zipfile.BadZipFile:
                raise ValueError(f"Excel文件损坏（ZIP格式无效）: {filepath}")
            except Exception as e:
                # 其他异常可能是权限问题等，记录警告但继续
                logger.warning(f"⚠️ ZIP验证时出现异常（可能不影响使用）: {e}")
        except Exception as e:
            if isinstance(e, ValueError):
                raise
            logger.error(f"验证ZIP格式失败: {filepath}, 错误: {e}")
            raise ValueError(f"验证Excel文件格式失败: {str(e)}")
    
    try:
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(_validate)
            try:
                future.result(timeout=timeout)
            except FutureTimeoutError:
                logger.warning(f"⚠️ ZIP格式验证超时（{timeout}秒），但继续处理: {filepath}")
                # ZIP验证超时不阻塞，只记录警告
                future.cancel()
    except Exception as e:
        if isinstance(e, ValueError):
            raise
        logger.warning(f"⚠️ ZIP格式验证异常，但继续处理: {e}")


def _validate_excel_structure(filepath: str, timeout: float = 2.0) -> None:
    """
    验证Excel文件结构完整性（带超时保护）
    
    参数:
        filepath: Excel文件路径
        timeout: 超时时间（秒），默认2秒
    
    异常:
        ValueError: Excel文件结构不完整
    """
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
    
    def _validate():
        """在后台线程中验证Excel结构"""
        try:
            with zipfile.ZipFile(filepath, 'r') as zf:
                namelist = zf.namelist()
                
                # 检查必需的Excel文件
                required_files = [
                    '[Content_Types].xml',
                    'xl/workbook.xml'
                ]
                
                missing_files = []
                for req_file in required_files:
                    # 检查文件是否存在（可能路径略有不同）
                    found = False
                    for name in namelist:
                        if name == req_file or name.endswith('/' + req_file):
                            found = True
                            break
                    if not found:
                        missing_files.append(req_file)
                
                if missing_files:
                    raise ValueError(
                        f"Excel文件结构不完整，缺少必需文件: {', '.join(missing_files)}: {filepath}"
                    )
        except zipfile.BadZipFile:
            # 如果ZIP文件本身有问题，这个应该在之前的检查中发现
            raise ValueError(f"Excel文件损坏（ZIP格式无效）: {filepath}")
        except Exception as e:
            if isinstance(e, ValueError):
                raise
            logger.error(f"验证Excel结构失败: {filepath}, 错误: {e}")
            raise ValueError(f"验证Excel文件结构失败: {str(e)}")
    
    try:
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(_validate)
            try:
                future.result(timeout=timeout)
            except FutureTimeoutError:
                logger.warning(f"⚠️ Excel结构验证超时（{timeout}秒），但继续处理: {filepath}")
                # 结构验证超时不阻塞，只记录警告
                future.cancel()
    except Exception as e:
        if isinstance(e, ValueError):
            raise
        logger.warning(f"⚠️ Excel结构验证异常，但继续处理: {e}")


def _validate_excel_row_count(filepath: str, sheet_name: str = None, max_rows: int = 10000, timeout: float = 5.0) -> None:
    """
    验证Excel文件的行数（带超时保护）
    
    参数:
        filepath: Excel文件路径
        sheet_name: 工作表名称（可选），如果为None则检查第一个工作表
        max_rows: 最大允许行数，默认10000
        timeout: 超时时间（秒），默认5秒
    
    异常:
        ValueError: 行数超过限制
    """
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
    
    def _check_rows():
        """在后台线程中检查行数"""
        try:
            file_ext = Path(filepath).suffix.lower()
            
            # 如果是 .xls 格式，需要先转换（但这里只做快速检查，不转换）
            # 对于 .xls 文件，跳过行数检查（因为转换需要时间）
            if file_ext == '.xls':
                logger.debug(f"⚠️ .xls 文件跳过行数检查（需要转换后才能检查）: {filepath}")
                return
            
            # 使用 read_only=True 模式，快速读取行数
            wb = load_workbook(filepath, data_only=True, read_only=True)
            
            try:
                # 选择工作表
                if sheet_name:
                    if sheet_name not in wb.sheetnames:
                        raise ValueError(f"工作表 '{sheet_name}' 不存在: {filepath}")
                    ws = wb[sheet_name]
                else:
                    if not wb.sheetnames:
                        raise ValueError(f"Excel文件不包含任何工作表: {filepath}")
                    ws = wb[wb.sheetnames[0]]  # 检查第一个工作表
                
                # 获取最大行数
                row_count = ws.max_row
                logger.info(f"📊 Excel文件行数检查: {row_count} 行（限制: {max_rows} 行）")
                
                if row_count > max_rows:
                    raise ValueError(
                        f"Excel文件行数过多: {row_count} 行，超过限制（{max_rows} 行）: {filepath}"
                    )
            finally:
                wb.close()
        except Exception as e:
            if isinstance(e, ValueError):
                raise
            logger.error(f"检查Excel行数失败: {filepath}, 错误: {e}")
            raise ValueError(f"检查Excel文件行数失败: {str(e)}")
    
    try:
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(_check_rows)
            try:
                future.result(timeout=timeout)
            except FutureTimeoutError:
                logger.warning(f"⚠️ Excel行数检查超时（{timeout}秒），但继续处理: {filepath}")
                # 行数检查超时不阻塞，只记录警告
                future.cancel()
    except Exception as e:
        if isinstance(e, ValueError):
            raise
        logger.warning(f"⚠️ Excel行数检查异常，但继续处理: {e}")


def print_excel_raw_data(filepath: str, sheet_name: str = None, max_rows: int = 15, max_cols: int = 25) -> None:
    """
    打印Excel文件的原始数据（前15行，前25列）
    
    参数:
        filepath: Excel文件路径
        sheet_name: 工作表名称（可选）
        max_rows: 最大行数，默认15
        max_cols: 最大列数，默认25
    """
    try:
        logger.info(f"📂 [DEBUG] print_excel_raw_data: 开始处理文件 {filepath}")
        from openpyxl import load_workbook
        
        # 加载工作簿
        logger.info(f"⏳ [DEBUG] print_excel_raw_data: 开始加载工作簿 (read_only=True)...")
        wb = load_workbook(filepath, data_only=True, read_only=True)
        logger.info(f"✅ [DEBUG] print_excel_raw_data: 工作簿加载完成")
        
        # 选择工作表
        logger.info(f"📋 [DEBUG] print_excel_raw_data: 选择工作表...")
        if sheet_name:
            ws = wb[sheet_name]
            logger.info(f"✅ [DEBUG] print_excel_raw_data: 使用指定工作表: {sheet_name}")
        else:
            if not wb.sheetnames:
                logger.warning("⚠️ [DEBUG] print_excel_raw_data: Excel文件不包含任何工作表")
                print("⚠️ Excel文件不包含任何工作表")
                return
            ws = wb[wb.sheetnames[0]]
            logger.info(f"✅ [DEBUG] print_excel_raw_data: 使用默认工作表: {ws.title}")
        
        # 确定实际读取范围
        logger.info(f"📏 [DEBUG] print_excel_raw_data: 工作表大小 - 最大行: {ws.max_row}, 最大列: {ws.max_column}")
        actual_max_col = min(ws.max_column, max_cols)
        actual_max_row = min(ws.max_row, max_rows)
        logger.info(f"📏 [DEBUG] print_excel_raw_data: 实际读取范围 - 行: {actual_max_row}, 列: {actual_max_col}")
        
        # 打印原始数据
        logger.info(f"🖨️ [DEBUG] print_excel_raw_data: 开始打印数据...")
        print("=" * 80)
        print(f"【最初传入的Excel原始数据 - 控制台打印】（前{actual_max_row}行，前{actual_max_col}列）")
        print(f"文件: {os.path.basename(filepath)}")
        print(f"工作表: {ws.title}")
        print("=" * 80)
        sys.stdout.flush()
        
        logger.info(f"🔄 [DEBUG] print_excel_raw_data: 开始遍历单元格...")
        for row in range(1, actual_max_row + 1):
            row_data = []
            for col in range(1, actual_max_col + 1):
                value = ws.cell(row, col).value
                row_data.append(value)
            print(f"行{row}: {row_data}")
            sys.stdout.flush()
            if row % 5 == 0:  # 每5行记录一次日志
                logger.info(f"📊 [DEBUG] print_excel_raw_data: 已处理 {row}/{actual_max_row} 行")
        
        print("=" * 80)
        sys.stdout.flush()
        logger.info(f"✅ [DEBUG] print_excel_raw_data: 数据打印完成")
        
        # 关闭工作簿
        logger.info(f"🔒 [DEBUG] print_excel_raw_data: 关闭工作簿...")
        print("🔍 [DEBUG] print_excel_raw_data: 准备关闭工作簿（使用print输出）")
        sys.stdout.flush()
        wb.close()
        print("🔍 [DEBUG] print_excel_raw_data: 工作簿已关闭（使用print输出）")
        sys.stdout.flush()
        logger.info(f"✅ [DEBUG] print_excel_raw_data: 工作簿已关闭")
        
        # 显式删除引用，帮助垃圾回收
        del wb
        del ws
        import gc
        gc.collect()  # 强制垃圾回收
        print("🔍 [DEBUG] print_excel_raw_data: 垃圾回收完成（使用print输出）")
        sys.stdout.flush()
        
        logger.info(f"🏁 [DEBUG] print_excel_raw_data: 函数即将返回，所有操作已完成")
        # 强制刷新输出
        sys.stdout.flush()
        logger.info(f"✅ [DEBUG] print_excel_raw_data: 函数执行完成，准备返回")
        # 使用 print 直接输出，确保能看到
        print("🔍 [DEBUG] print_excel_raw_data: 函数即将返回（使用print输出）")
        sys.stdout.flush()
        # 最后一条日志
        logger.info(f"🏁 [DEBUG] print_excel_raw_data: 函数返回前最后一条日志")
        print("🔍 [DEBUG] print_excel_raw_data: 函数返回（使用print输出）")
        sys.stdout.flush()
        return  # 显式返回
        
    except Exception as e:
        logger.error(f"❌ [DEBUG] print_excel_raw_data: 打印Excel原始数据失败: {filepath}, 错误: {e}", exc_info=True)
        print(f"⚠️ 打印Excel原始数据失败: {str(e)}")
        sys.stdout.flush()


def _get_preview_data_lightweight(filepath: str, sheet_name: str = None, max_rows: int = 15, max_cols: int = 25, timeout: int = 10, max_file_size_mb: Optional[int] = None, max_excel_rows: Optional[int] = None) -> Tuple[List[List[Any]], int]:
    """
    轻量级获取Excel预览数据（使用read_only模式，用于表头分析）
    
    参数:
        filepath: Excel文件路径
        sheet_name: 工作表名称（可选）
        max_rows: 最大行数，默认15
        max_cols: 最大列数，默认25
        timeout: 超时时间（秒），默认10秒
        max_file_size_mb: 最大文件大小（MB），如果为None则使用默认值
    
    返回:
        (预览数据列表, 最大列数)
    """
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
    
    def _read_preview():
        """在后台线程中读取预览数据"""
        try:
            # 文件预检查（基础检查 + ZIP验证）
            _validate_excel_file_basic(filepath, max_file_size_mb=max_file_size_mb)
            file_ext = Path(filepath).suffix.lower()
            if file_ext == '.xlsx':
                _validate_xlsx_format(filepath, timeout=0.5)
                # 行数检查（超过配置的最大行数直接拒绝）
                try:
                    max_rows_value = max_excel_rows if max_excel_rows is not None else 10000
                    _validate_excel_row_count(filepath, sheet_name=sheet_name, max_rows=max_rows_value, timeout=5.0)
                except ValueError as row_error:
                    # 捕获行数检查异常，转换为可识别的异常
                    if "行数过多" in str(row_error) or "超过限制" in str(row_error):
                        # 重新抛出，但标记为行数限制错误
                        raise ValueError(f"Excel文件行数超过限制: {str(row_error)}") from row_error
                    raise
            
            # 使用 read_only=True 模式，更快且更轻量
            wb = load_workbook(filepath, data_only=True, read_only=True)
            
            # 选择工作表
            if sheet_name:
                ws = wb[sheet_name]
            else:
                if not wb.sheetnames:
                    raise ValueError("Excel文件不包含任何工作表")
                ws = wb[wb.sheetnames[0]]
            
            # 确定实际读取范围
            actual_max_col = min(ws.max_column, max_cols)
            actual_max_row = min(ws.max_row, max_rows)
            max_col = ws.max_column  # 保存总列数
            
            # 读取数据
            data = []
            for row in range(1, actual_max_row + 1):
                row_data = []
                for col in range(1, actual_max_col + 1):
                    value = ws.cell(row, col).value
                    row_data.append(value)
                data.append(row_data)
            
            wb.close()
            return data, max_col
        except Exception as e:
            logger.error(f"读取Excel预览数据失败: {filepath}, 错误: {e}")
            raise
    
    try:
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(_read_preview)
            try:
                data, max_col = future.result(timeout=timeout)
                return data, max_col
            except FutureTimeoutError:
                logger.error(f"读取Excel预览数据超时: {filepath} (超时时间: {timeout}秒)")
                future.cancel()
                raise TimeoutError(f"读取Excel预览数据超时（{timeout}秒）: {filepath}")
    except Exception as e:
        if isinstance(e, TimeoutError):
            raise
        # 如果是行数检查异常，重新抛出以便上层处理
        if isinstance(e, ValueError) and ("行数超过限制" in str(e) or "行数过多" in str(e)):
            raise
        logger.error(f"读取Excel预览数据时发生异常: {filepath}, 错误: {e}")
        raise


def _analyze_header_with_llm_lightweight(preview_data: List[List[Any]], max_col: int,
                                         llm_api_key: Optional[str] = None,
                                         llm_base_url: Optional[str] = None,
                                         llm_model: Optional[str] = None,
                                         timeout: Optional[int] = None,
                                         thinking_callback: Optional[callable] = None) -> Tuple[HeaderAnalysis, str]:
    """
    使用LLM分析表头结构（轻量级版本，不需要SmartHeaderProcessor）
    
    参数:
        preview_data: 预览数据列表
        max_col: 最大列数
        llm_api_key: LLM API密钥（可选）
        llm_base_url: LLM API地址（可选）
        llm_model: LLM模型名称（可选）
        timeout: 超时时间（秒），默认90秒
        thinking_callback: 用于流式输出 thinking 内容的回调函数（可选）
    
    返回:
        (分析结果, LLM原始响应)
    """
    # 构建提示词
    num_cols = len(preview_data[0]) if preview_data else 0
    num_rows = len(preview_data)
    
    # 构建简单的表格字符串
    table_str = "【Excel原始数据】（前15行，前25列）\n\n"
    table_str += "行号 | " + " | ".join([f"列{i+1}" for i in range(num_cols)]) + "\n"
    table_str += "-" * (8 + num_cols * 15) + "\n"
    
    for i, row in enumerate(preview_data, 1):
        row_str = " | ".join([str(cell) if cell is not None else "" for cell in row])
        table_str += f"  {i:2d}  | {row_str}\n"
    
    prompt = f"""你是一个Excel表格结构分析专家。请分析以下Excel表格的原始数据，识别表头结构。

{table_str}

【总列数】{max_col}

## 分析任务

请分析表格结构，识别：
1. **无效行（skip_rows）**：表头之前的无效行（如文档标题、说明文字、注释、公司名称、填报说明等）
2. **表头行数（header_rows）**：所有表头行，包括多级表头的所有层级
3. **表头类型（header_type）**：single（单表头）或 multi（多级表头）
4. **数据起始行（data_start_row）**：数据开始的行号，必须等于 skip_rows + header_rows + 1
5. **数据起始列（start_col）**：第一个表头行中第一个非空表头开始的列号

## 识别规则

### 无效行特征：
- 文档标题（如"2024年度报表"）
- 公司名称或部门名称（如"XX公司"、"XX部门"）
- 填报说明（如"填报机构"、"填报日期"、"填报机构/日期"等，任何包含"填报"关键词的行）
- 只有数字没有标签的行（如只有"222"、"111"等数字，没有对应的列名）
- 完全空行或只有少量文本的行

### 表头行特征：
- 包含列名或分类标签（如"销售事业部"、"华东大区"、"线上销售额"等）
- 有明确的层级结构（多级表头）
- 通常不包含大量数值数据

### 数据行特征：
- 包含大量数值数据
- 不再是表头文本或分类标签

## 输出格式

请以JSON格式返回分析结果：

```json
{{
    "skip_rows": <表头之前的无效行数，如果第1行就是表头则填0>,
    "header_rows": <表头占用的总行数>,
    "header_type": "<single或multi>",
    "data_start_row": <数据开始行号（1-indexed），必须等于skip_rows+header_rows+1>,
    "start_col": <数据起始列号（1-indexed）>,
    "valid_cols": null,
    "confidence": "<high/medium/low>",
    "reason": "<详细说明识别过程>"
}}
```

## 注意事项

1. 行号和列号都从1开始计数
2. data_start_row 必须等于 skip_rows + header_rows + 1
3. valid_cols 始终设为 null
4. 只返回JSON，不要其他内容
5. 如果第1行就是表头，则 skip_rows=0
6. 多级表头的所有行都要计入 header_rows

只返回JSON，不要其他内容。"""
    
    # 调用LLM
    from .config import EXCEL_LLM_API_KEY, EXCEL_LLM_BASE_URL, EXCEL_LLM_MODEL
    api_key = llm_api_key if llm_api_key is not None else EXCEL_LLM_API_KEY
    if not api_key:
        raise ValueError("LLM API密钥未配置，无法进行Excel分析。请配置EXCEL_LLM_API_KEY或传入llm_api_key参数")
    
    # 调用LLM API
    result = _call_llm_api(prompt, api_key, llm_base_url or EXCEL_LLM_BASE_URL, llm_model or EXCEL_LLM_MODEL, timeout=timeout, thinking_callback=thinking_callback)
    
    if not result:
        raise ValueError("LLM分析失败：无法获取LLM响应，请检查API配置")
    
    # 解析LLM分析结果
    analysis = _parse_llm_analysis_response_lightweight(result)
    
    return analysis, result


def _call_llm_api(prompt: str, 
                  llm_api_key: str,
                  llm_base_url: Optional[str] = None,
                  llm_model: Optional[str] = None,
                  timeout: Optional[int] = None,
                  thinking_callback: Optional[callable] = None) -> str:
    """
    调用LLM API（独立函数，不依赖SmartHeaderProcessor）
    """
    from .config import EXCEL_LLM_BASE_URL, EXCEL_LLM_MODEL
    
    base_url = llm_base_url or EXCEL_LLM_BASE_URL
    model = llm_model or EXCEL_LLM_MODEL
    
    if not base_url:
        raise ValueError("LLM API地址未配置，请配置EXCEL_LLM_BASE_URL或传入llm_base_url参数")
    
    if not model:
        raise ValueError("LLM模型名称未配置，请配置EXCEL_LLM_MODEL或传入llm_model参数")
    
    # 使用传入的超时时间，默认90秒
    request_timeout = timeout if timeout is not None else 90
    
    logger.info(f"⏱️ 超时设置: {request_timeout} 秒")
    
    # 构建请求URL和参数
    # 注意：base_url 应该已经包含完整的路径（如 /v1/chat/completions），直接使用
    url = base_url
    headers = {
        "Authorization": f"Bearer {llm_api_key}",
        "Content-Type": "application/json"
    }
    
    # 构建消息
    messages = [{"role": "user", "content": prompt}]
    
    # 使用流式调用以支持 thinking 功能（默认启用）
    base_payload = {
        "model": model,
        "messages": messages,
        "temperature": 0.4,
        "max_tokens": 1000,
        "stream": True,  # 流式调用（必须启用以支持 thinking）
    }
    
    # 提供默认回调函数
    if thinking_callback is None:
        def default_callback(content: str):
            pass  # 空回调，不执行任何操作
        thinking_callback = default_callback
    
    logger.info(f"📡 发送 LLM API 请求到: {url} (流式调用)")
    logger.info(f"📝 提示词长度: {len(prompt)} 字符")
    
    try:
        # 默认启用 thinking 功能（流式输出）
        payload_with_thinking = base_payload.copy()
        payload_with_thinking["enable_thinking"] = True  # 默认启用 thinking
        
        logger.info("💭 已启用 Thinking 功能，将实时流式输出思考过程")
        
        response = requests.post(
            url, 
            headers=headers, 
            json=payload_with_thinking, 
            timeout=request_timeout,
            stream=True  # 启用流式响应
        )
        
        # 如果启用 thinking 失败，回退到不使用 thinking
        if response.status_code != 200:
            try:
                error_json = response.json()
                if "enable_thinking" in str(error_json).lower():
                    logger.warning("⚠️ API 不支持 enable_thinking 参数，将回退到不使用 thinking")
                    logger.warning("💭 注意：Thinking 流式输出将不可用（API 不支持）")
                    payload_no_thinking = base_payload.copy()
                    response = requests.post(
                        url, 
                        headers=headers, 
                        json=payload_no_thinking, 
                        timeout=request_timeout,
                        stream=True
                    )
            except:
                pass
        
        # 如果请求失败，输出详细的错误信息
        if response.status_code != 200:
            error_detail = ""
            try:
                # 对于流式响应，尝试读取错误信息
                error_text = ""
                for line in response.iter_lines():
                    if line:
                        error_text += line.decode('utf-8') + "\n"
                error_detail = error_text
            except:
                try:
                    error_json = response.json()
                    error_detail = json.dumps(error_json, ensure_ascii=False, indent=2)
                except:
                    error_detail = response.text
            
            logger.error(f"❌ LLM API 调用失败 (状态码: {response.status_code})")
            logger.error(f"📋 错误详情: {error_detail}")
            response.raise_for_status()
        
        # 处理流式响应
        full_content = ""
        for line in response.iter_lines():
            if line:
                line_str = line.decode('utf-8')
                if line_str.startswith('data: '):
                    data_str = line_str[6:]
                    if data_str.strip() == '[DONE]':
                        break
                    try:
                        chunk_data = json.loads(data_str)
                        if 'choices' in chunk_data and len(chunk_data['choices']) > 0:
                            delta = chunk_data['choices'][0].get('delta', {})
                            # 使用 get 方法并提供默认值，避免 None 值
                            content = delta.get('content', '')
                            if content:
                                full_content += content
                                thinking_callback(content)
                    except json.JSONDecodeError:
                        continue
        
        if not full_content:
            logger.warning("⚠️ LLM 流式响应为空")
            return None
        
        logger.info("✅ LLM API 调用成功")
        return full_content
    except requests.exceptions.RequestException as e:
        logger.error(f"❌ LLM调用失败 (网络错误): {e}")
        if hasattr(e, 'response') and e.response is not None:
            try:
                error_json = e.response.json()
                logger.error(f"📋 API 错误响应: {json.dumps(error_json, ensure_ascii=False, indent=2)}")
            except:
                logger.error(f"📋 API 错误响应 (文本): {e.response.text}")
        logger.debug("异常详情:", exc_info=True)
        return None
    except Exception as e:
        logger.error(f"❌ LLM调用失败: {e}")
        logger.debug("异常详情:", exc_info=True)
        return None


def _parse_llm_analysis_response_lightweight(response: str) -> HeaderAnalysis:
    """解析LLM分析结果（轻量级版本，不依赖SmartHeaderProcessor）"""
    if not response:
        raise ValueError("LLM响应为空")
    
    try:
        # 提取JSON部分（支持嵌套JSON）
        start_idx = response.find('{')
        end_idx = response.rfind('}')
        if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
            json_str = response[start_idx:end_idx + 1]
            data = json.loads(json_str)
        else:
            # 如果找不到完整的JSON，尝试用正则匹配
            json_match = re.search(r'\{.*\}', response, re.DOTALL)
            if not json_match:
                raise ValueError("未找到JSON格式的响应")
            data = json.loads(json_match.group())
        
        # 解析有效列（始终为None）
        valid_cols = None
        
        # 解析起始列（默认为1）
        start_col = int(data.get('start_col', 1))
        if start_col < 1:
            start_col = 1
        
        # 构建HeaderAnalysis对象
        analysis = HeaderAnalysis(
            skip_rows=int(data.get('skip_rows', 0)),
            header_rows=int(data.get('header_rows', 1)),
            header_type=data.get('header_type', 'single'),
            data_start_row=int(data.get('data_start_row', 1)),
            start_col=start_col,
            confidence=data.get('confidence', 'medium'),
            reason=f"LLM分析: {data.get('reason', '')}",
            valid_cols=valid_cols
        )
        
        logger.info(f"✅ LLM分析完成:")
        logger.info(f"  - 跳过行数: {analysis.skip_rows}")
        logger.info(f"  - 表头行数: {analysis.header_rows}")
        logger.info(f"  - 表头类型: {analysis.header_type}")
        logger.info(f"  - 数据起始行: {analysis.data_start_row}")
        logger.info(f"  - 数据起始列: {analysis.start_col}")
        logger.info(f"  - 置信度: {analysis.confidence}")
        
        return analysis
    except (json.JSONDecodeError, KeyError, ValueError) as e:
        logger.error(f"❌ 解析LLM分析响应失败: {e}")
        logger.error(f"📋 响应内容: {response[:500]}")
        raise ValueError(f"解析LLM分析响应失败: {e}")


def _save_csv_with_timeout(df: pd.DataFrame, csv_path: str, timeout: int = 30) -> None:
    """带超时保护的保存CSV文件"""
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
    
    def _save():
        """在后台线程中保存CSV"""
        try:
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        except Exception as e:
            logger.error(f"保存CSV文件失败: {csv_path}, 错误: {e}")
            raise
    
    try:
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(_save)
            try:
                future.result(timeout=timeout)
            except FutureTimeoutError:
                logger.error(f"保存CSV文件超时: {csv_path} (超时时间: {timeout}秒)")
                future.cancel()
                raise TimeoutError(f"保存CSV文件超时（{timeout}秒）: {csv_path}")
    except Exception as e:
        if isinstance(e, TimeoutError):
            raise
        logger.error(f"保存CSV文件时发生异常: {csv_path}, 错误: {e}")
        raise


def process_excel_file(
    filepath: str,
    output_dir: str,
    sheet_name: str = None,
    use_llm_validate: bool = False,  # 已废弃，保留用于兼容性，现在总是使用LLM
    output_filename: str = None,
    llm_api_key: Optional[str] = None,
    llm_base_url: Optional[str] = None,
    llm_model: Optional[str] = None,
    preprocessing_timeout: Optional[int] = None,
    excel_processing_timeout: Optional[int] = None,  # Excel处理超时时间（秒），在LLM分析之前
    debug_print_header_analysis: bool = False,  # 是否流式打印原始数据（用于调试）
    thinking_callback: Optional[callable] = None,  # 用于流式输出 thinking 内容的回调函数
    max_file_size_mb: Optional[int] = None,  # 最大文件大小（MB），如果为None则使用默认值
    max_rows: Optional[int] = None  # 最大行数，如果为None则使用默认值10000
) -> ExcelProcessResult:
    """
    处理Excel文件的主函数
    
    参数:
        filepath: Excel文件路径
        output_dir: 输出目录
        sheet_name: 工作表名称
        use_llm_validate: 已废弃，保留用于兼容性。现在总是使用LLM进行分析
        output_filename: 输出文件名（不含扩展名）
        llm_api_key: LLM API密钥（必填，否则会抛出异常）
        llm_base_url: LLM API地址（可选）
        llm_model: LLM模型名称（可选）
        preprocessing_timeout: 预处理超时时间（秒），默认90秒
        excel_processing_timeout: Excel处理超时时间（秒），默认10秒（包括文件加载和数据读取）
        debug_print_header_analysis: 是否流式打印原始数据（用于调试），默认False
        max_file_size_mb: 最大文件大小（MB），如果为None则使用默认值
    
    返回:
        ExcelProcessResult
    
    注意:
        现在必须使用LLM进行分析，不再支持规则分析。请确保提供llm_api_key参数。
    """
    try:
        logger.info(f"🚀 [DEBUG] process_excel_file: 开始处理文件 {filepath}")
        # 确保输出目录存在
        logger.info(f"📁 [DEBUG] process_excel_file: 确保输出目录存在: {output_dir}")
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"✅ [DEBUG] process_excel_file: 输出目录已准备")
        
        # 设置超时时间
        excel_processing_timeout_seconds = excel_processing_timeout if excel_processing_timeout is not None else 10
        logger.info(f"⏱️ [DEBUG] process_excel_file: Excel处理超时时间: {excel_processing_timeout_seconds}秒")
        
        # 第一步：使用轻量级方式获取预览数据并进行LLM分析（不需要创建SmartHeaderProcessor）
        logger.info(f"📂 [DEBUG] process_excel_file: 开始获取预览数据（轻量级模式）")
        logger.info(f"📂 [DEBUG] process_excel_file: 文件路径: {filepath}, 工作表: {sheet_name}")
        
        try:
            preview_data, max_col = _get_preview_data_lightweight(
                filepath, 
                sheet_name, 
                max_rows=15, 
                max_cols=25, 
                timeout=excel_processing_timeout_seconds,
                max_file_size_mb=max_file_size_mb,
                max_excel_rows=max_rows
            )
            logger.info(f"✅ [DEBUG] process_excel_file: 预览数据获取完成，共 {len(preview_data)} 行，{max_col} 列")
        except TimeoutError as e:
            error_msg = f"获取Excel预览数据超时: {str(e)}"
            logger.error(f"❌ {error_msg}")
            return ExcelProcessResult(
                success=False,
                header_analysis=None,
                processed_file_path=None,
                metadata_file_path=None,
                column_names=[],
                column_metadata={},
                row_count=0,
                error_message=error_msg
            )
        except Exception as e:
            # 捕获所有异常，检查是否是行数检查异常（因为可能是从线程池抛出的）
            error_str = str(e)
            error_type = type(e).__name__
            
            # 检查是否是行数相关的错误（支持多种格式）
            is_row_limit_error = (
                "行数超过限制" in error_str or 
                "行数过多" in error_str or 
                ("超过限制" in error_str and "行" in error_str) or
                (isinstance(e, ValueError) and "行数" in error_str and "限制" in error_str)
            )
            
            if is_row_limit_error:
                # 提取行数信息
                import re
                match = re.search(r'(\d+)\s*行', error_str)
                row_count = match.group(1) if match else "未知"
                match_limit = re.search(r'限制[（(](\d+)\s*行', error_str)
                if not match_limit:
                    match_limit = re.search(r'超过限制[（(](\d+)\s*行', error_str)
                limit = match_limit.group(1) if match_limit else (max_rows if max_rows else 10000)
                error_msg = f"Excel文件行数过多（{row_count} 行），超过限制（{limit} 行）。请减少文件行数或调整配置中的最大行数限制。"
                logger.warning(f"⚠️ {error_msg}")
                return ExcelProcessResult(
                    success=False,
                    header_analysis=None,
                    processed_file_path=None,
                    metadata_file_path=None,
                    column_names=[],
                    column_metadata={},
                    row_count=0,
                    error_message=error_msg
                )
            # 其他异常继续抛出
            logger.error(f"❌ 获取Excel预览数据时发生未处理的异常: {error_type}: {error_str}")
            raise
        
        # 第二步：使用预览数据进行LLM分析
        logger.info("🤖 开始LLM表头分析（使用预览数据）...")
        try:
            analysis, llm_response = _analyze_header_with_llm_lightweight(
                preview_data,
                max_col,
                llm_api_key=llm_api_key,
                llm_base_url=llm_base_url,
                llm_model=llm_model,
                timeout=preprocessing_timeout,
                thinking_callback=thinking_callback
            )
            logger.info("✅ LLM表头分析完成")
        except Exception as e:
            error_msg = f"LLM表头分析失败: {str(e)}"
            logger.error(f"❌ {error_msg}")
            return ExcelProcessResult(
                success=False,
                header_analysis=None,
                processed_file_path=None,
                metadata_file_path=None,
                column_names=[],
                column_metadata={},
                row_count=0,
                error_message=error_msg
            )
        
        # 第三步：创建SmartHeaderProcessor来读取完整数据（只有在需要读取完整数据时才创建）
        logger.info(f"📂 [DEBUG] process_excel_file: 开始创建 SmartHeaderProcessor（用于读取完整数据）")
        load_timeout = excel_processing_timeout_seconds
        read_timeout = excel_processing_timeout_seconds
        
        from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
        
        def _create_processor():
            """在后台线程中创建 SmartHeaderProcessor"""
            try:
                return SmartHeaderProcessor(
                    filepath, 
                    sheet_name, 
                    load_timeout=load_timeout, 
                    read_timeout=read_timeout,
                    debug_print_header_analysis=debug_print_header_analysis,
                    max_file_size_mb=max_file_size_mb,
                    max_rows=max_rows
                )
            except Exception as e:
                logger.error(f"创建 SmartHeaderProcessor 失败: {e}")
                raise
        
        try:
            # 使用总超时时间（excel_processing_timeout_seconds）来保护整个初始化过程
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(_create_processor)
                try:
                    processor = future.result(timeout=excel_processing_timeout_seconds)
                    logger.info("✅ [DEBUG] process_excel_file: SmartHeaderProcessor 创建完成")
                except FutureTimeoutError:
                    logger.error(f"创建 SmartHeaderProcessor 超时: {filepath} (超时时间: {excel_processing_timeout_seconds}秒)")
                    future.cancel()
                    error_msg = f"Excel文件处理超时（{excel_processing_timeout_seconds}秒）: {filepath}"
                    logger.error(f"❌ {error_msg}")
                    return ExcelProcessResult(
                        success=False,
                        header_analysis=analysis,  # 保留已完成的LLM分析结果
                        processed_file_path=None,
                        metadata_file_path=None,
                        column_names=[],
                        column_metadata={},
                        row_count=0,
                        error_message=error_msg
                    )
        except TimeoutError as e:
            error_msg = f"Excel文件加载超时: {str(e)}"
            logger.error(f"❌ {error_msg}")
            return ExcelProcessResult(
                success=False,
                header_analysis=analysis,  # 保留已完成的LLM分析结果
                processed_file_path=None,
                metadata_file_path=None,
                column_names=[],
                column_metadata={},
                row_count=0,
                error_message=error_msg
            )
        
        # 第四步：使用SmartHeaderProcessor读取完整数据并转换为DataFrame
        logger.info("📊 开始读取完整数据并转换为DataFrame...")
        df, _, column_metadata, _ = processor.to_dataframe(
            analysis=analysis,  # 使用已完成的LLM分析结果
            use_llm_validate=False,  # 不再需要LLM分析，因为已经完成了
            llm_api_key=llm_api_key,
            llm_base_url=llm_base_url,
            llm_model=llm_model,
            preprocessing_timeout=preprocessing_timeout,
            thinking_callback=thinking_callback
        )
        logger.info("✅ 完整数据读取完成")
        processor.close()
        
        # 生成输出文件名
        if not output_filename:
            base_name = Path(filepath).stem
            output_filename = f"{base_name}_processed"
        
        # 保存CSV（带超时保护）
        csv_path = os.path.join(output_dir, f"{output_filename}.csv")
        try:
            _save_csv_with_timeout(df, csv_path, timeout=30)
        except TimeoutError as e:
            logger.error(f"保存CSV文件超时: {csv_path}")
            raise ValueError(f"保存CSV文件超时: {str(e)}")
        
        # 提取字段值样本（分组聚合后的常见值）
        logger.info("📊 提取字段值样本...")
        column_value_samples = extract_column_value_samples(df, max_samples_per_column=10)
        
        # 将值样本信息合并到列元数据中
        for col_name, samples in column_value_samples.items():
            if col_name in column_metadata:
                column_metadata[col_name]["value_samples"] = samples
            else:
                # 如果列不在元数据中（理论上不应该发生），创建新的元数据项
                column_metadata[col_name] = {"value_samples": samples}
        
        # 保存元数据
        metadata = {
            "header_analysis": analysis.to_dict(),
            "column_metadata": column_metadata,
            "column_names": list(df.columns),
            "row_count": len(df),
            "original_file": os.path.basename(filepath)
        }
        metadata_path = os.path.join(output_dir, f"{output_filename}_metadata.json")
        with open(metadata_path, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)
        
        # 打印处理后的JSON元数据（暂时注释）
        # logger.info("=" * 80)
        # logger.info("📄 处理后的JSON元数据:")
        # logger.info("=" * 80)
        # logger.info(json.dumps(metadata, ensure_ascii=False, indent=2))
        # logger.info("=" * 80)
        
        return ExcelProcessResult(
            success=True,
            header_analysis=analysis,
            processed_file_path=csv_path,
            metadata_file_path=metadata_path,
            column_names=list(df.columns),
            column_metadata=column_metadata,
            row_count=len(df),
            error_message=None,
            llm_analysis_response=llm_response
        )
        
    except Exception as e:
        import traceback
        error_msg = f"{str(e)}\n{traceback.format_exc()}"
        return ExcelProcessResult(
            success=False,
            header_analysis=None,
            processed_file_path=None,
            metadata_file_path=None,
            column_names=[],
            column_metadata={},
            row_count=0,
            error_message=error_msg
        )


def get_sheet_names(filepath: str, timeout: int = 10, max_file_size_mb: Optional[int] = None, max_rows: Optional[int] = None) -> List[str]:
    """获取Excel文件的所有工作表名称（带超时保护）
    
    Args:
        filepath: Excel文件路径
        timeout: 超时时间（秒），默认10秒
        max_file_size_mb: 最大文件大小（MB），如果为None则使用默认值
    
    Returns:
        工作表名称列表，如果超时或出错则返回空列表
    """
    import threading
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
    
    logger.info(f"📋 [DEBUG] get_sheet_names: 开始获取工作表名称，文件: {os.path.basename(filepath)}, 超时: {timeout}秒")
    print(f"🔍 [DEBUG] get_sheet_names: 开始获取工作表名称（使用print输出）")
    sys.stdout.flush()
    
    def _load_sheets():
        """在后台线程中加载工作表名称"""
        try:
            # 文件预检查（基础检查 + ZIP验证）
            _validate_excel_file_basic(filepath, max_file_size_mb=max_file_size_mb)
            file_ext = Path(filepath).suffix.lower()
            if file_ext == '.xlsx':
                _validate_xlsx_format(filepath, timeout=0.5)
                # 行数检查（超过配置的最大行数直接拒绝）
                try:
                    max_rows_value = max_rows if max_rows is not None else 10000
                    _validate_excel_row_count(filepath, sheet_name=None, max_rows=max_rows_value, timeout=5.0)
                except ValueError as row_error:
                    # 捕获行数检查异常，记录警告但不抛出（get_sheet_names 返回空列表表示失败）
                    if "行数过多" in str(row_error) or "超过限制" in str(row_error):
                        logger.warning(f"⚠️ Excel文件行数超过限制，无法获取工作表列表: {str(row_error)}")
                        return []  # 返回空列表表示失败
                    raise
            
            logger.info(f"📂 [DEBUG] get_sheet_names: 开始加载工作簿...")
            print(f"🔍 [DEBUG] get_sheet_names: 开始加载工作簿（使用print输出）")
            sys.stdout.flush()
            wb = load_workbook(filepath)
            logger.info(f"✅ [DEBUG] get_sheet_names: 工作簿加载完成")
            print(f"🔍 [DEBUG] get_sheet_names: 工作簿加载完成（使用print输出）")
            sys.stdout.flush()
            sheets = wb.sheetnames
            logger.info(f"📋 [DEBUG] get_sheet_names: 获取到工作表: {sheets}")
            wb.close()
            logger.info(f"✅ [DEBUG] get_sheet_names: 工作簿已关闭")
            return sheets
        except Exception as e:
            logger.warning(f"❌ [DEBUG] get_sheet_names: 读取Excel工作表失败: {filepath}, 错误: {e}", exc_info=True)
            return []
    
    try:
        # 使用线程池执行，带超时保护
        logger.info(f"🔄 [DEBUG] get_sheet_names: 准备在线程池中执行...")
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(_load_sheets)
            try:
                logger.info(f"⏳ [DEBUG] get_sheet_names: 等待结果，超时时间: {timeout}秒...")
                sheets = future.result(timeout=timeout)
                logger.info(f"✅ [DEBUG] get_sheet_names: 获取工作表名称成功: {sheets}")
                return sheets if sheets else []
            except FutureTimeoutError:
                logger.error(f"获取Excel工作表名称超时: {filepath} (超时时间: {timeout}秒)")
                # 尝试取消任务（但可能已经执行了）
                future.cancel()
                return []
    except Exception as e:
        logger.error(f"获取Excel工作表名称时发生异常: {filepath}, 错误: {e}")
        return []


def extract_column_value_samples(
    df: pd.DataFrame,
    max_samples_per_column: int = 10,
    max_unique_ratio: float = 0.5
) -> Dict[str, Dict[str, Any]]:
    """
    提取每个字段的常见值样本（通过分组聚合）
    
    参数:
        df: 数据框
        max_samples_per_column: 每个字段最多保留的样本数量
        max_unique_ratio: 如果唯一值占比超过此比例，则只提供统计信息而不统计频率
    
    返回:
        字典，key为列名，value为包含常见值和统计信息的字典
    """
    column_samples = {}
    
    for col_name in df.columns:
        col_data = df[col_name]
        
        # 跳过完全为空的列
        if col_data.isna().all():
            continue
        
        # 计算非空值数量
        non_null_count = col_data.notna().sum()
        if non_null_count == 0:
            continue
        
        # 计算唯一值数量
        unique_count = col_data.nunique()
        unique_ratio = unique_count / non_null_count if non_null_count > 0 else 1.0
        
        sample_info = {
            "total_count": len(col_data),
            "non_null_count": int(non_null_count),
            "null_count": int(col_data.isna().sum()),
            "unique_count": int(unique_count),
            "data_type": str(col_data.dtype)
        }
        
        # 判断是否为数值类型
        is_numeric = pd.api.types.is_numeric_dtype(col_data)
        
        if is_numeric:
            # 数值类型：提供统计信息和常见值（如果唯一值不太多）
            sample_info["is_numeric"] = True
            non_null_data = col_data.dropna()
            if len(non_null_data) > 0:
                sample_info["min"] = float(non_null_data.min())
                sample_info["max"] = float(non_null_data.max())
                sample_info["mean"] = float(non_null_data.mean())
                sample_info["median"] = float(non_null_data.median())
            else:
                sample_info["min"] = None
                sample_info["max"] = None
                sample_info["mean"] = None
                sample_info["median"] = None
            
            # 如果唯一值不太多，也统计频率
            if unique_ratio <= max_unique_ratio and unique_count <= 100:
                value_counts = col_data.value_counts().head(max_samples_per_column)
                sample_info["top_values"] = [
                    {"value": float(k) if pd.notna(k) else None, "count": int(v)}
                    for k, v in value_counts.items()
                ]
            elif unique_count <= max_samples_per_column:
                # 即使唯一值比例高，但如果总数不多，也展示所有值
                value_counts = col_data.value_counts().head(max_samples_per_column)
                sample_info["top_values"] = [
                    {"value": float(k) if pd.notna(k) else None, "count": int(v)}
                    for k, v in value_counts.items()
                ]
                sample_info["note"] = f"唯一值较多（{unique_count}个），展示所有值"
        else:
            # 非数值类型：统计频率
            sample_info["is_numeric"] = False
            
            # 如果唯一值太多，只提供统计信息
            if unique_ratio > max_unique_ratio:
                sample_info["note"] = f"唯一值较多（{unique_count}个），仅展示部分常见值"
                # 仍然展示前N个最常见的值
                value_counts = col_data.value_counts().head(max_samples_per_column)
                sample_info["top_values"] = [
                    {"value": str(k) if pd.notna(k) else "空值", "count": int(v)}
                    for k, v in value_counts.items()
                ]
            else:
                # 唯一值不太多，统计所有值的频率
                value_counts = col_data.value_counts().head(max_samples_per_column)
                sample_info["top_values"] = [
                    {"value": str(k) if pd.notna(k) else "空值", "count": int(v)}
                    for k, v in value_counts.items()
                ]
        
        column_samples[col_name] = sample_info
    
    return column_samples


def _build_column_hierarchy_tree(column_metadata: Dict[str, Dict]) -> str:
    """
    构建列层级结构的树形展示
    
    参数:
        column_metadata: 列元数据字典
    
    返回:
        格式化的树形结构字符串
    """
    if not column_metadata:
        return ""
    
    # 构建树形结构
    tree = {}
    
    for col_name, meta in column_metadata.items():
        # 获取所有层级
        levels = []
        level_keys = sorted([k for k in meta.keys() if k.startswith('level')], 
                          key=lambda x: int(x.replace('level', '')))
        for level_key in level_keys:
            value = meta.get(level_key)
            if value and str(value).strip():
                levels.append(str(value).strip())
        
        # 如果没有层级信息，使用列名本身
        if not levels:
            levels = [col_name]
        
        # 构建树
        current = tree
        for i, level_value in enumerate(levels):
            if level_value not in current:
                current[level_value] = {}
            current = current[level_value]
    
    # 递归生成树形字符串
    def _format_tree(node: Dict, prefix: str = "", is_last: bool = True, depth: int = 0) -> List[str]:
        lines = []
        items = list(node.items())
        
        for idx, (key, children) in enumerate(items):
            is_last_item = (idx == len(items) - 1)
            current_prefix = "└─ " if is_last_item else "├─ "
            
            if children:
                # 有子节点
                lines.append(f"{prefix}{current_prefix}{key}")
                next_prefix = prefix + ("   " if is_last_item else "│  ")
                child_lines = _format_tree(children, next_prefix, is_last_item, depth + 1)
                lines.extend(child_lines)
            else:
                # 叶子节点
                lines.append(f"{prefix}{current_prefix}{key}")
        
        return lines
    
    tree_lines = _format_tree(tree)
    return "\n".join(tree_lines)


def generate_analysis_prompt(
    process_result: ExcelProcessResult,
    custom_prompt: str = None,
    include_metadata: bool = True
) -> str:
    """
    根据Excel处理结果生成数据分析提示词
    
    参数:
        process_result: Excel处理结果
        custom_prompt: 自定义分析提示词
        include_metadata: 是否包含列结构元数据
    
    返回:
        格式化的提示词
    """
    if not process_result.success:
        return ""
    
    # 基础信息
    prompt_parts = []
    
    # 添加语言要求（必须在最前面）
    prompt_parts.append("**重要要求：请使用中文进行所有分析和回答，包括代码注释、分析报告等所有内容。**")
    prompt_parts.append("")
    prompt_parts.append("**禁止要求：请不要生成任何图表绘制代码，包括但不限于：**")
    prompt_parts.append("- 不要使用 matplotlib、plotly、seaborn 等绘图库")
    prompt_parts.append("- 不要使用 plt.figure()、plt.plot()、plt.savefig() 等绘图函数")
    prompt_parts.append("- 不要使用 .plot()、.hist() 等 pandas 绘图方法")
    prompt_parts.append("- 不要保存任何图片文件（.png、.jpg、.svg 等）")
    prompt_parts.append("**请专注于数据分析和统计计算，不要生成可视化代码。**")
    prompt_parts.append("")
    
    if custom_prompt:
        prompt_parts.append(custom_prompt)
    else:
        prompt_parts.append("请对上传的数据进行全面分析，生成数据分析报告。")
    
    # 添加数据文件信息（重要：告诉AI需要读取CSV文件）
    if process_result.processed_file_path:
        csv_filename = os.path.basename(process_result.processed_file_path)
        prompt_parts.append(f"\n\n## 数据文件")
        prompt_parts.append(f"**重要：工作空间中已准备好处理后的CSV数据文件，文件名为：`{csv_filename}`**")
        prompt_parts.append(f"")
        prompt_parts.append(f"**请务必使用以下代码读取数据文件进行分析：**")
        prompt_parts.append(f"```python")
        prompt_parts.append(f"import pandas as pd")
        prompt_parts.append(f"")
        prompt_parts.append(f"# 读取处理后的CSV文件")
        prompt_parts.append(f"df = pd.read_csv('{csv_filename}')")
        prompt_parts.append(f"print(f'数据形状: {{df.shape}}')")
        prompt_parts.append(f"print(f'列名: {{list(df.columns)}}')")
        prompt_parts.append(f"```")
        prompt_parts.append(f"")
        prompt_parts.append(f"**注意：**")
        prompt_parts.append(f"- CSV文件已保存在当前工作空间目录中")
        prompt_parts.append(f"- 请使用 `pd.read_csv('{csv_filename}')` 读取数据")
        prompt_parts.append(f"- 不要仅根据元数据进行分析，必须读取实际数据文件进行计算")
        prompt_parts.append(f"")
    
    # 添加数据概况
    prompt_parts.append(f"\n## 数据概况")
    prompt_parts.append(f"- 数据行数: {process_result.row_count}")
    prompt_parts.append(f"- 列数: {len(process_result.column_names)}")
    
    # 添加表头类型信息（仅保留对分析有用的信息）
    if process_result.header_analysis:
        ha = process_result.header_analysis
        if ha.header_type == 'multi':
            prompt_parts.append(f"\n## 表头结构")
            prompt_parts.append(f"- 表头类型: 多级表头（{ha.header_rows}层）")
    
    # 添加列结构元数据（帮助AI理解列之间的关系）
    if include_metadata and process_result.column_metadata:
        # 检查是否有多级结构
        has_multi_level = any(
            len(meta) > 1 
            for meta in process_result.column_metadata.values()
        )
        
        if has_multi_level:
            prompt_parts.append(f"\n## 列层级结构（多级表头语义关系）")
            prompt_parts.append("以下树形结构展示了列之间的层级分组关系，有助于理解数据的业务含义：")
            prompt_parts.append("")
            hierarchy_tree = _build_column_hierarchy_tree(process_result.column_metadata)
            if hierarchy_tree:
                prompt_parts.append(hierarchy_tree)
            else:
                # 如果树形构建失败，使用分组展示
                groups = defaultdict(list)
                for col_name, meta in process_result.column_metadata.items():
                    level1 = meta.get('level1', col_name)
                    groups[level1].append(col_name)
                
                for group, cols in groups.items():
                    if len(cols) > 1:
                        prompt_parts.append(f"- {group}: {', '.join(cols)}")
    
    # 添加完整的列名列表
    prompt_parts.append(f"\n## 完整列名列表")
    if len(process_result.column_names) <= 30:
        # 如果列数不多，全部展示
        for idx, col_name in enumerate(process_result.column_names, 1):
            prompt_parts.append(f"{idx}. {col_name}")
    else:
        # 如果列数很多，展示前20个和后10个
        for idx, col_name in enumerate(process_result.column_names[:20], 1):
            prompt_parts.append(f"{idx}. {col_name}")
        prompt_parts.append(f"... (省略中间 {len(process_result.column_names) - 30} 列) ...")
        for idx, col_name in enumerate(process_result.column_names[-10:], len(process_result.column_names) - 9):
            prompt_parts.append(f"{idx}. {col_name}")
        prompt_parts.append(f"\n(共 {len(process_result.column_names)} 列)")
    
    # 添加字段值样本信息（以JSON格式提供，更结构化）
    if include_metadata and process_result.column_metadata:
        prompt_parts.append(f"\n## 字段值样本（常见值统计）")
        prompt_parts.append("以下JSON格式展示了每个字段的常见值及其出现频率，有助于理解数据的实际内容：")
        prompt_parts.append("")
        
        # 构建包含值样本的column_metadata JSON
        column_metadata_with_samples = {}
        for col_name in process_result.column_names:
            if col_name in process_result.column_metadata:
                column_metadata_with_samples[col_name] = process_result.column_metadata[col_name]
        
        # 将column_metadata转换为格式化的JSON字符串
        prompt_parts.append("```json")
        prompt_parts.append(json.dumps(column_metadata_with_samples, ensure_ascii=False, indent=2))
        prompt_parts.append("```")
        prompt_parts.append("")
        
        prompt_parts.append("**说明：**")
        prompt_parts.append("- 每个字段的元数据包含 `value_samples` 字段，其中包含该字段的统计信息和常见值")
        prompt_parts.append("- `value_samples.top_values` 数组展示了出现频率最高的值及其出现次数")
        prompt_parts.append("- 对于数值类型字段，还包含 `min`、`max`、`mean`、`median` 等统计信息")
    
    # 在末尾再次强调要求
    prompt_parts.append("\n\n**再次提醒：请务必使用中文进行所有分析、代码注释和报告撰写，且不要生成任何图表绘制代码。**")
    
    full_prompt = '\n'.join(prompt_parts)
    
    # 打印生成的提示词
    logger.info("=" * 80)
    logger.info("📝 生成的AI分析提示词:")
    logger.info("=" * 80)
    logger.info(full_prompt)
    logger.info("=" * 80)
    
    return full_prompt

